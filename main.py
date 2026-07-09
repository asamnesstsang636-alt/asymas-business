import streamlit as st
import pandas as pd
st.set_page_config(
    page_title="ASYMAS BUSINESS",
    page_icon="🌾",
    layout="wide",
    initial_sidebar_state="auto"
)
st.markdown("""
<meta name="mobile-web-app-capable" content="yes">
""", unsafe_allow_html=True)
from supabase import create_client, Client
from datetime import date, datetime, timedelta
from fpdf import FPDF
import base64
import io
import qrcode
from PIL import Image
import tempfile
import os
import json
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from streamlit_qrcode_scanner import qrcode_scanner

# === CONFIG SUPABASE ===
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# === FONCTIONS ===
@st.cache_data(ttl=60)
def load_table(table_name):
    try:
        with st.spinner(f"Chargement {table_name}..."):
            data = supabase.table(table_name).select("*").order("id", desc=True).execute()
        return pd.DataFrame(data.data)
    except Exception as e:
        st.error(f"Erreur chargement {table_name}")
        st.code(repr(e))
        return pd.DataFrame()

@st.cache_data(ttl=300)
def get_table_columns(table_name):
    try:
        test = supabase.table(table_name).select("*").limit(1).execute()
        if test.data:
            return list(test.data[0].keys())
        return []
    except:
        return []

@st.cache_data(ttl=10)
def load_passwords():
    try:
        data = supabase.table("utilisateurs").select("nom,role,password,permissions,categories_autorisees").execute()
        passwords = {}
        perms = {}
        for user in data.data:
            key = user['nom'].strip().upper() # ✅ ON UTILISE NOM EN MAJUSCULE
            passwords[key] = user['password']
            perms[key] = {
                'permissions': user.get('permissions', {}),
                'categories_autorisees': user.get('categories_autorisees', []),
                'role': user.get('role')
            }
        st.session_state.permissions_db = perms
        return passwords
    except Exception as e:
        st.warning("Mode hors ligne: Utilisation des mots de passe par défaut")
        st.session_state.permissions_db = {}
        return {
            "TSANG": "tsang2024", # ✅ CLÉS = NOMS
            "ASIYA": "asiya2024",
            "BASAM": "basam2024"
        }

def generer_qrcode(data_text):
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=3, border=1)
    qr.add_data(data_text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
    img.save(temp_file.name)
    return temp_file.name

def safe_pdf_txt(txt):
    if txt is None or pd.isna(txt):
        return ""
    txt = str(txt)
    txt = txt.replace('—', '-').replace('–', '-').replace('’', "'").replace('“', '"').replace('”', '"')
    txt = txt.replace('•', '-').replace('…', '...')
    txt = ''.join(c if ord(c) < 128 else '?' for c in txt)
    return txt.replace('\n', ' ').replace('\r', '').strip()

def generer_pdf_facture(numero, type_op, client, details_list, montant, devise, tel_client="+243...", periode="", type_facture="Simple"):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=False, margin=10)
    pdf.set_fill_color(20, 50, 40)
    pdf.rect(0, 0, 210, 35, 'F')
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", "B", 20)
    pdf.set_xy(10, 8)
    pdf.cell(0, 10, "ASYMAS BUSINESS", ln=True)
    pdf.set_font("Arial", "", 9)
    pdf.set_xy(10, 16)
    pdf.cell(0, 5, "Beni, Nord-Kivu, RDC | Tel: +243 995 105 623", ln=True)
    pdf.set_xy(10, 21)
    pdf.cell(0, 5, "Email: asamnesstsang636@gmail.com", ln=True)
    pdf.set_font("Arial", "B", 10)
    pdf.set_xy(150, 8)
    titre_fact = "FACTURE N" if type_facture == "Simple" else "PROFORMA N"
    pdf.cell(50, 6, titre_fact, ln=True, align="R")
    pdf.set_font("Arial", "", 10)
    pdf.set_xy(150, 14)
    pdf.cell(50, 6, safe_pdf_txt(numero), ln=True, align="R")
    pdf.set_font("Arial", "", 9)
    pdf.set_xy(150, 20)
    pdf.cell(50, 6, f"Date: {date.today().strftime('%d/%m/%Y')}", ln=True, align="R")
    y_pos = 45
    pdf.set_text_color(0, 0, 0)
    pdf.set_fill_color(255, 204, 0)
    pdf.set_font("Arial", "B", 14)
    pdf.set_xy(10, y_pos)
    pdf.cell(0, 10, f"{type_facture.upper()} {safe_pdf_txt(type_op.upper())}", ln=True, fill=True)
    y_pos += 15
    pdf.set_font("Arial", "B", 10)
    pdf.set_draw_color(0, 0, 0)
    pdf.set_xy(10, y_pos)
    pdf.cell(85, 7, "FACTURE A:", 1, 0, 'L')
    pdf.cell(10, 7, "", 0, 0)
    pdf.cell(85, 7, "DETAILS PAIEMENT:", 1, 1, 'L')
    y_pos += 7
    pdf.set_font("Arial", "", 9)
    pdf.set_xy(10, y_pos)
    pdf.cell(85, 6, f"Client: {safe_pdf_txt(client)}", 'LR', 0, 'L')
    pdf.cell(10, 6, "", 0, 0)
    pdf.cell(85, 6, "M-Pesa: +243817264448", 'LR', 1, 'L')
    y_pos += 6
    pdf.set_xy(10, y_pos)
    pdf.cell(85, 6, f"Tel: {safe_pdf_txt(tel_client)}", 'LR', 0, 'L')
    pdf.cell(10, 6, "", 0, 0)
    pdf.cell(85, 6, "Echeance: Immediate", 'LR', 1, 'L')
    y_pos += 6
    pdf.set_xy(10, y_pos)
    pdf.cell(85, 6, f"Date emission: {date.today().strftime('%d/%m/%Y')}", 'LRB', 0, 'L')
    pdf.cell(10, 6, "", 0, 0)
    pdf.cell(85, 6, "", 'LRB', 1, 'L')
    y_pos += 14
    pdf.set_fill_color(0, 102, 0)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", "B", 10)
    pdf.set_xy(10, y_pos)
    pdf.cell(115, 8, "DESIGNATION", 1, 0, 'C', True)
    pdf.cell(25, 8, "QTE", 1, 0, 'C', True)
    pdf.cell(40, 8, f"MONTANT ({safe_pdf_txt(devise)})", 1, 1, 'C', True)
    y_pos += 8
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 9)
    if isinstance(details_list, list) and details_list:
        for item in details_list:
            if y_pos > 240:
                pdf.add_page()
                y_pos = 30
            nom = safe_pdf_txt(item.get('nom', ''))
            qte = item.get('qte', 1)
            pu = item.get('pu', item.get('prix', 0))
            montant_item = pu * qte
            pdf.set_xy(10, y_pos)
            pdf.cell(115, 7, nom, 1, 0, 'L')
            pdf.cell(25, 7, str(qte), 1, 0, 'C')
            pdf.cell(40, 7, f"{montant_item:,.0f}", 1, 1, 'R')
            y_pos += 7
    else:
        pdf.set_xy(10, y_pos)
        pdf.cell(115, 7, safe_pdf_txt(details_list), 1, 0, 'L')
        pdf.cell(25, 7, "1", 1, 0, 'C')
        pdf.cell(40, 7, f"{montant:,.0f}", 1, 1, 'R')
        y_pos += 7
    if periode:
        if y_pos > 240:
            pdf.add_page()
            y_pos = 30
        pdf.set_xy(10, y_pos)
        pdf.cell(115, 7, f"Periode: {safe_pdf_txt(periode)}", 1, 0, 'L')
        pdf.cell(25, 7, "", 1, 0, 'C')
        pdf.cell(40, 7, "", 1, 1, 'R')
        y_pos += 7
    pdf.set_fill_color(255, 204, 0)
    pdf.set_font("Arial", "B", 11)
    pdf.set_xy(10, y_pos)
    pdf.cell(140, 10, "MONTANT TOTAL A PAYER", 1, 0, 'R', True)
    pdf.cell(40, 10, f"{montant:,.0f} {safe_pdf_txt(devise)}", 1, 1, 'R', True)
    y_pos += 15
    if y_pos > 220:
        pdf.add_page()
        y_pos = 30
    pdf.set_xy(10, y_pos)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 8, "SIGNATURE RESPONSABLE:", ln=True)
    y_pos += 11
    pdf.set_draw_color(0, 0, 0)
    pdf.line(10, y_pos, 100, y_pos)
    y_pos += 1
    pdf.set_font("Arial", "", 9)
    pdf.set_xy(10, y_pos)
    pdf.cell(90, 5, "Ing. SAMY TSANGYA", ln=True)
    y_pos += 5
    pdf.set_xy(10, y_pos)
    pdf.cell(90, 5, "Tel: +243 995 105 623", ln=True)
    y_pos += 5
    pdf.set_xy(10, y_pos)
    pdf.cell(90, 5, "Beni, Nord-Kivu, RDC", ln=True)
    y_pos += 10
    pdf.set_font("Arial", "I", 10)
    pdf.set_text_color(0, 102, 0)
    pdf.set_xy(10, y_pos)
    pdf.cell(0, 6, "Merci pour votre confiance! ASYMAS BUSINESS - Votre partenaire de croissance", ln=True, align="C")
    qr_data = f"""ASYMAS BUSINESS
Facture: {numero}
Type: {type_op}
Client: {client}
Montant: {montant:,.0f} {devise}
Date: {date.today().strftime('%d/%m/%Y')}
Tel: +243 995 105 623"""
    qr_path = generer_qrcode(qr_data)
    pdf.image(qr_path, x=155, y=y_pos-25, w=25)
    os.unlink(qr_path)
    return bytes(pdf.output(dest='S'))

#... GARDE TES AUTRES FONCTIONS PDF ET EXCEL ICI...

st.markdown("""
<link rel="manifest" href="data:application/manifest+json,{
  \"name\": \"ASYMAS BUSINESS\",
  \"short_name\": \"ASYMAS\",
  \"start_url\": \".\",
  \"display\": \"standalone\",
  \"background_color\": \"#000000\",
  \"theme_color\": \"#00ff41\",
  \"description\": \"Agriculture Commerce Immobilier Automobile\",
  \"icons\": [{
    \"src\": \"https://raw.githubusercontent.com/twitter/twemoji/master/assets/72x72/1f48e.png\",
    \"sizes\": \"192x192\",
    \"type\": \"image/png\"
  }]
}">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-capable" content="yes">
""", unsafe_allow_html=True)

st.markdown("""
<style>
#MainMenu {visibility: hidden!important;}
header {visibility: hidden!important;}
.stAppToolbar {display: none!important;}
[data-testid="stToolbar"] {display: none!important;}
[data-testid="stDecoration"] {display: none!important;}
[data-testid="stHeader"] {display: none!important;}
footer {visibility: hidden!important;}
.stDeployButton {display:none!important;}
[data-testid="stStatusWidget"] {display: none!important;}
[data-testid="manage-app-button"] {display: none!important;}
iframe[src*="streamlit.io"] {display: none!important;}
button[kind="header"] {display: none!important;}
div[data-testid="stBottomBlockContainer"] {display: none!important;}
.st-emotion-cache-1wbqy5l {display: none!important;}
button[title="Manage app"] {display: none!important;}
a[href*="share.streamlit.io"] {display: none!important;}
</style>
""", unsafe_allow_html=True)

# === SYSTEME DE CONNEXION ===
passwords_db = load_passwords()

if 'user_role' not in st.session_state:
    st.session_state.user_role = None
    st.session_state.user_name = None
    st.session_state.user_perms = {}
    st.session_state.user_cats = []

if st.session_state.user_role is None:
    st.markdown("# 🔐 ASYMAS BUSINESS - CONNEXION")
    col1, col2, col3 = st.columns([1,2,1])
    with col2:
        st.markdown("### Connectez-vous :")

        nom_connect = st.text_input("Nom d'utilisateur", placeholder="", value="")
        password = st.text_input("Mot de passe", type="password", placeholder="", value="", key="pwd")

        if st.button("SE CONNECTER", width="stretch", type="primary"):
            if nom_connect and password:
                try:
                    nom_key = nom_connect.strip().upper()

                    # ✅ AJOUT DE password DANS LE SELECT
                    response = supabase.table("utilisateurs")\
             .select("id, nom, role, password, permissions, categories_autorisees")\
             .ilike("nom", nom_connect.strip())\
             .execute()

                    df_users_login = pd.DataFrame(response.data)

                    if not df_users_login.empty:
                        user = df_users_login.iloc[0]
                        mdp_attendu = user['password'] # ✅ ON PREND LE MDP DE LA DB DIRECT

                        if mdp_attendu and password == mdp_attendu:
                            st.session_state.user_role = user['role']
                            st.session_state.user_name = user['nom']
                            st.session_state.user_perms = user.get('permissions', {})
                            st.session_state.user_cats = user.get('categories_autorisees', [])
                            st.success(f"Bienvenue {st.session_state.user_name}")
                            st.rerun()
                        else:
                            st.error("Nom d'utilisateur ou mot de passe incorrect")
                    else:
                        st.error("Nom d'utilisateur non trouvé")

                except Exception as e:
                    st.error(f"❌ ERREUR CONNEXION: {repr(e)}") # ✅ AFFICHE L'ERREUR RÉELLE
                    st.code(str(e))
            else:
                st.warning("Veuillez remplir tous les champs")
    st.stop()

# === SIDEBAR APRÈS CONNEXION ===
if st.session_state.user_role is not None:
    with st.sidebar:
        st.success(f"👤 {st.session_state.user_name}")
        st.caption(f"Rôle: {st.session_state.user_role}")
        if 'theme_choisi' not in st.session_state:
            st.session_state.theme_choisi = "Sombre ASYMAS"
        theme = st.selectbox("🎨 Thème", ["Sombre ASYMAS","Bleu Pro","Vert Agri","Noir Luxe"], key="theme_choisi")
        if st.button("🚪 Déconnexion", width="stretch"):
            st.session_state.user_role=None
            st.session_state.user_name=None
            st.session_state.user_perms={}
            st.session_state.user_cats=[]
            st.rerun()
    if theme=="Sombre ASYMAS": st.markdown("""<style>.stApp{background:#0E1117;color:#E0E0E0}h1,h2,h3{color:#14B814!important}</style>""",unsafe_allow_html=True)
    elif theme=="Bleu Pro": st.markdown("""<style>.stApp{background:#0A1929;color:#E3F2FD}h1,h2,h3{color:#2196F3!important}</style>""",unsafe_allow_html=True)
    elif theme=="Vert Agri": st.markdown("""<style>.stApp{background:#1B2A1B;color:#E8F5E9}h1,h2,h3{color:#4CAF50!important}</style>""",unsafe_allow_html=True)
    elif theme=="Noir Luxe": st.markdown("""<style>.stApp{background:#000;color:#FFF}h1,h2,h3{color:#FFD700!important}</style>""",unsafe_allow_html=True)

st.markdown("""
<style>
h1, h2, h3 {
    color: #00ff41!important;
    font-size: 2.2rem!important;
    font-weight: 900!important;
    padding: 10px 0!important;
    border-bottom: 3px solid #00ff41!important;
    margin-bottom: 20px!important;
}
div[data-testid="stMetricValue"] {color: #00ff41!important;}
.stButton>button {background-color: #00ff41!important; color: black!important; font-weight: bold; border: none;}
</style>
""", unsafe_allow_html=True)

df_biens = load_table("biens")
df_articles = load_table("articles")
df_voitures = load_table("voitures")
df_compta = load_table("compta")
df_factures = load_table("factures_proforma")
df_devis = load_table("devis")
df_utilisateurs = load_table("utilisateurs")

if 'montant' not in df_compta.columns:
    df_compta['montant'] = 0
if 'type' not in df_compta.columns:
    df_compta['type'] = 'Inconnu'
if 'date' in df_compta.columns:
    df_compta['date'] = pd.to_datetime(df_compta['date'], errors='coerce')
    df_compta = df_compta.sort_values('date', ascending=False)

st.markdown(f"# ASYMAS BUSINESS - {st.session_state.user_name}")
st.markdown("### Agriculture • Commerce • Immobilier • Automobile • Beni RDC")

with st.sidebar:
    st.markdown(f"## 👤 {st.session_state.user_name}")
    st.markdown(f"**Rôle : {st.session_state.user_role}**")
    st.info("ASYMAS BUSINESS v2.6")
    if st.button("🔄 Actualiser", key="btn_save"):
        st.cache_data.clear()
        st.rerun()

perms = st.session_state.user_perms
if isinstance(perms, str):
    try: perms = json.loads(perms)
    except: perms = {}

tabs_dispo = []
if st.session_state.user_role == "PDG" or perms.get('dashboard', True):
    tabs_dispo.append("📊 Dashboard")
if st.session_state.user_role == "PDG" or perms.get('commerce', True):
    tabs_dispo.append("🛍️ Commerce")
if st.session_state.user_role == "PDG" or perms.get('stock', False):
    tabs_dispo.append("📦 Gestion Stock")
if st.session_state.user_role == "PDG" or perms.get('immobilier', False):
    tabs_dispo.append("🏠 Immobilier")
if st.session_state.user_role == "PDG" or perms.get('automobile', False):
    tabs_dispo.append("🚗 Automobile")
if st.session_state.user_role == "PDG" or perms.get('parc', False):
    tabs_dispo.append("🚘 Gestion Parc")
if st.session_state.user_role == "PDG" or perms.get('comptabilite', False):
    tabs_dispo.append("💰 Comptabilité")
if st.session_state.user_role == "PDG" or perms.get('factures', False):
    tabs_dispo.append("📄 Factures")
if st.session_state.user_role == "PDG" or perms.get('devis_industriel', False) or perms.get('devis_batiment', False):
    tabs_dispo.append("📋 Devis")
if st.session_state.user_role == "PDG" or perms.get('users', False):
    tabs_dispo.append("👥 Utilisateurs")

if not tabs_dispo:
    tabs_dispo = ["📊 Dashboard", "🛍️ Commerce"]

tabs = st.tabs(tabs_dispo)
tab_map = {name: tab for name, tab in zip(tabs_dispo, tabs)}

if "📊 Dashboard" in tab_map:
    with tab_map["📊 Dashboard"]:
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("🏠 Biens", len(df_biens))
        col2.metric("📦 Articles", len(df_articles))
        col3.metric("🚗 Voitures", len(df_voitures))
        if not df_compta.empty and 'type' in df_compta.columns and 'montant' in df_compta.columns:
            revenus = df_compta[df_compta['type']=='Revenu']['montant'].sum()
            col4.metric("💰 Revenus", f"{revenus:,.0f} FC")
        elif not df_compta.empty:
            col4.metric("💰 Écritures", len(df_compta))
        else:
            col4.metric("💰 Revenus", "0 FC")

if "🛍️ Commerce" in tab_map:
    with tab_map["🛍️ Commerce"]:
        st.markdown("## 🛍️ Commerce - Point de Vente")
        if 'panier_commerce' not in st.session_state:
            st.session_state.panier_commerce = []
        if 'vente_finie' not in st.session_state:
            st.session_state.vente_finie = False
        if 'pdf_data' not in st.session_state:
            st.session_state.pdf_data = None
        if 'num_fact' not in st.session_state:
            st.session_state.num_fact = None
        if 'client_com_nom' not in st.session_state:
            st.session_state.client_com_nom = ""
        if 'client_com_tel' not in st.session_state:
            st.session_state.client_com_tel = "+243..."
        if 'last_qr' not in st.session_state:
            st.session_state.last_qr = ""

        col_gauche, col_droite = st.columns([2,1])
        with col_gauche:
            st.subheader("👤 Client")
            st.session_state.client_com_nom = st.text_input("Nom Client", value=st.session_state.client_com_nom, key="nom_client_c")
            st.session_state.client_com_tel = st.text_input("Téléphone Client", value=st.session_state.client_com_tel, key="tel_client_c")
            st.subheader("🔍 Scanner QR Code")
            col_scan1, col_scan2 = st.columns([2,1])
            with col_scan1:
                qr_code = qrcode_scanner(key='qr_commerce_unique')
            with col_scan2:
                recherche_manuelle = st.text_input("🔎 Recherche manuelle", placeholder="Tape le nom...", key="search_man_c")
            if qr_code and qr_code!= st.session_state.last_qr:
                st.session_state.last_qr = qr_code
                st.rerun()

            df_articles_filtre = df_articles[df_articles['stock'] > 0].copy()
            if qr_code:
                qr_clean = str(qr_code).strip().upper()
                df_articles_filtre = df_articles_filtre[df_articles_filtre['code_qr'].astype(str).str.strip().str.upper() == qr_clean]
                if not df_articles_filtre.empty:
                    st.success(f"✅ QR Trouvé : {df_articles_filtre.iloc[0]['nom_article']}")
                else:
                    st.error(f"❌ QR {qr_code} : Produit introuvable")
            elif recherche_manuelle:
                mask = df_articles_filtre['nom_article'].str.contains(recherche_manuelle, case=False, na=False)
                df_articles_filtre = df_articles_filtre[mask]

            if df_articles_filtre.empty:
                st.warning("⚠️ Aucun produit disponible")
            else:
                st.success(f"✅ {len(df_articles_filtre)} produit(s) disponible(s)")
                options_articles = []
                for _, p in df_articles_filtre.iterrows():
                    qr_txt = f" | QR:{p['code_qr']}" if 'code_qr' in p and p['code_qr'] else ""
                    prix_usd = f" | {p['prix_vente_usd']:,.2f}$" if 'prix_vente_usd' in p else ""
                    options_articles.append(f"{p['nom_article']} | Stock:{int(p['stock'])} | {p['prix_vente']:,.0f} FC{prix_usd}{qr_txt} | ID:{p['id']}")
                article_choisi = st.selectbox("Sélectionne le produit", options_articles, key="select_article_unique")
                if article_choisi:
                    id_choisi = int(article_choisi.split("ID:")[1])
                    p = df_articles_filtre[df_articles_filtre['id'] == id_choisi].iloc[0]
                    c1, c2, c3 = st.columns(3)
                    qte_max = int(p['stock'])
                    qte = c1.number_input("Quantité", min_value=1, max_value=qte_max, value=1, key="qte_c_unique")
                    c2.metric("Stock dispo", qte_max)
                    c3.metric("Prix unitaire", f"{p['prix_vente']:,.0f} FC")
                    st.info(f"**{p['nom_article']}** | Catégorie: {p.get('categorie','N/A')} | QR: {p.get('code_qr','N/A')}")
                    if st.button("🛒 AJOUTER AU PANIER", type="primary", width="stretch", key="add_article_unique"):
                        existant = next((item for item in st.session_state.panier_commerce if item['id'] == int(p['id'])), None)
                        if existant:
                            if existant['qte'] + qte <= qte_max:
                                existant['qte'] += qte
                                st.success(f"Panier mis à jour: {existant['qte']}x")
                            else:
                                st.error(f"Stock insuffisant! Max dispo: {qte_max}")
                        else:
                            st.session_state.panier_commerce.append({
                                "id": int(p['id']),
                                "nom": str(p['nom_article']),
                                "pu": float(p['prix_vente']),
                                "qte": int(qte),
                                "code_qr": p.get('code_qr',''),
                                "stock_max": qte_max
                            })
                            st.success("Ajouté au panier")
                        st.rerun()
        with col_droite:
            st.subheader("🛒 Panier")
            if st.session_state.vente_finie and st.session_state.pdf_data:
                st.success("✅ Vente enregistrée!")
                st.download_button(
                    "📥 Télécharger Facture PDF",
                    data=st.session_state.pdf_data,
                    file_name=f"{st.session_state.num_fact}.pdf",
                    mime="application/pdf",
                    width="stretch"
                )
                pdf_b64 = base64.b64encode(st.session_state.pdf_data).decode()
                st.components.v1.html(f"""
                    <button onclick="printPDF()" style="width:100%; padding:10px; background:#00ff41; color:black; font-weight:bold; border:none; border-radius:5px; cursor:pointer; margin-top:10px;">
                        🖨️ IMPRIMER LA FACTURE
                    </button>
                    <script>
                    function printPDF() {{
                        const pdfData = 'data:application/pdf;base64,{pdf_b64}';
                        const win = window.open('', '_blank');
                        win.document.write('<iframe src="' + pdfData + '" width="100%" height="100%" style="border:none;"></iframe>');
                        win.document.close();
                        setTimeout(() => {{ win.print(); }}, 1000);
                    }}
                    </script>
                """, height=60)
                if st.button("NOUVELLE VENTE", width="stretch"):
                    st.session_state.vente_finie = False
                    st.session_state.pdf_data = None
                    st.session_state.num_fact = None
                    st.session_state.client_com_nom = ""
                    st.session_state.last_qr = ""
                    st.rerun()
            elif not st.session_state.panier_commerce:
                st.info("Panier vide")
            else:
                total_panier = 0
                for i, item in enumerate(st.session_state.panier_commerce):
                    col1, col2, col3 = st.columns([4,2,1])
                    col1.write(f"**{item['nom']}**")
                    col2.write(f"Qté: {item['qte']} | {item['pu']:,.0f} FC")
                    if col3.button("❌", key=f"d_{i}"):
                        st.session_state.panier_commerce.pop(i)
                        st.rerun()
                    total_panier += item['qte'] * item['pu']
                st.markdown(f"### Total: {total_panier:,.0f} FC")
                st.divider()
                if st.button("💾 FINALISER VENTE & FACTURE", width="stretch", type="primary"):
                    if not st.session_state.client_com_nom:
                        st.error("Nom du client obligatoire!")
                    else:
                        try:
                            num_fact = f"VTE-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                            details_list = []
                            for item in st.session_state.panier_commerce:
                                supabase.table("ventes").insert({
                                    "numero_facture": num_fact,
                                    "client_nom": st.session_state.client_com_nom,
                                    "article_id": item['id'],
                                    "quantite": item['qte'],
                                    "prix_unitaire": item['pu'],
                                    "total": item['qte'] * item['pu']
                                }).execute()
                                stock_actuel = df_articles[df_articles['id'] == item['id']]['stock'].iloc[0]
                                supabase.table("articles").update({"stock": int(stock_actuel - item['qte'])}).eq("id", item['id']).execute()
                                details_list.append({
                                    "nom": item['nom'],
                                    "qte": item['qte'],
                                    "pu": item['pu'],
                                    "total": item['qte'] * item['pu']
                                })
                            details_json = json.dumps(details_list)
                            supabase.table("compta").insert({
                                "date": str(date.today()),
                                "type": "Revenu",
                                "categorie": "Vente Commerce",
                                "description": f"Vente - {st.session_state.client_com_nom}",
                                "montant": float(total_panier),
                                "devise": "FC",
                                "numero_facture": num_fact,
                                "details": details_json,
                                "utilisateur": st.session_state.user_name
                            }).execute()
                            pdf_bytes = generer_pdf_facture(
                                num_fact, "Vente Commerce", st.session_state.client_com_nom,
                                details_list, total_panier, "FC", st.session_state.client_com_tel
                            )
                            st.session_state.pdf_data = pdf_bytes
                            st.session_state.num_fact = num_fact
                            st.session_state.vente_finie = True
                            st.session_state.panier_commerce = []
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error("Erreur finalisation vente")
                            st.code(repr(e))

if "📦 Gestion Stock" in tab_map:
    with tab_map["📦 Gestion Stock"]:
        st.markdown("## 📦 Gestion Stock Commerce - Articles & Pertes")
        
        tab_stock, tab_ajout, tab_mvt, tab_pertes = st.tabs(["📊 Stock Actuel", "➕ Ajouter Article", "📈 Mouvements", "⚠️ Pertes & Casses"])

        with tab_stock:
            st.subheader("📊 Stock Actuel Commerce")
            if df_articles.empty:
                st.info("Aucun article en stock")
            else:
                for _, row in df_articles.iterrows():
                    col1, col2, col3, col4 = st.columns([3,1,1,1])
                    with col1:
                        st.write(f"**{row['nom_article']}** - {row.get('categorie','')} - QR:{row.get('code_qr','N/A')}")
                    with col2:
                        stock_val = int(row.get('stock',0))
                        if stock_val < 5:
                            st.error(f"⚠️ Stock: {stock_val}")
                        else:
                            st.success(f"✅ Stock: {stock_val}")
                    with col3:
                        st.write(f"PA: {row.get('prix_achat',0):,.0f}")
                    with col4:
                        st.write(f"PV: {row.get('prix_vente',0):,.0f} FC")
                    
                    with st.expander(f"Modifier/Supprimer {row['nom_article']}"):
                        c1, c2, c3 = st.columns(3)
                        with c1:
                            new_nom = st.text_input("Nom", value=row['nom_article'], key=f"nom_art_{row['id']}")
                            new_cat = st.text_input("Catégorie", value=row.get('categorie',''), key=f"cat_art_{row['id']}")
                            new_code_qr = st.text_input("Code QR", value=row.get('code_qr',''), key=f"qr_art_{row['id']}")
                        with c2:
                            new_prix_a = st.number_input("Prix Achat FC", value=float(row.get('prix_achat',0)), key=f"pa_art_{row['id']}")
                            new_prix_v = st.number_input("Prix Vente FC", value=float(row.get('prix_vente',0)), key=f"pv_art_{row['id']}")
                            new_prix_usd = st.number_input("Prix Vente $", value=float(row.get('prix_vente_usd',0)), key=f"pusd_art_{row['id']}")
                        with c3:
                            new_stock = st.number_input("Stock", value=int(row.get('stock',0)), key=f"stock_art_{row['id']}")
                        
                        c1, c2 = st.columns(2)
                        if c1.button("✏️ Modifier", key=f"mod_art_{row['id']}", width="stretch"):
                            try:
                                data_update = {
                                    "nom_article": str(new_nom),
                                    "categorie": str(new_cat),
                                    "prix_achat": float(new_prix_a),
                                    "prix_vente": float(new_prix_v),
                                    "stock": int(new_stock),
                                    "code_qr": str(new_code_qr) if new_code_qr else None
                                }
                                colonnes_articles = get_table_columns("articles")
                                if "prix_vente_usd" in colonnes_articles:
                                    data_update["prix_vente_usd"] = float(new_prix_usd)
                                supabase.table("articles").update(data_update).eq("id", int(row['id'])).execute()
                                st.success("Modifié")
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as e:
                                st.error("Erreur modif")
                                st.code(repr(e))
                        if st.session_state.user_role == "PDG" or perms.get('supprimer', False):
                            if c2.button("🗑️ Supprimer", key=f"del_art_{row['id']}", width="stretch"):
                                try:
                                    supabase.table("articles").delete().eq("id", int(row['id'])).execute()
                                    st.success("Supprimé")
                                    st.cache_data.clear()
                                    st.rerun()
                                except Exception as e:
                                    st.error("Erreur suppression")
                                    st.code(repr(e))

        with tab_ajout:
            st.subheader("➕ Ajouter Nouvel Article Commerce")
            qr_scan_ajout = qrcode_scanner(key='qr_add_article_com')
            if qr_scan_ajout:
                st.success(f"QR scanné : {qr_scan_ajout}")
                st.session_state.qr_code_temp = qr_scan_ajout

            with st.form("form_article_com", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                nom = c1.text_input("Nom Article")
                cat = c2.text_input("Catégorie")
                code_qr = c3.text_input("Code QR", value=st.session_state.get('qr_code_temp', ''))
                c1, c2, c3 = st.columns(3)
                prix_achat_fc = c1.number_input("Prix Achat FC", min_value=0.0)
                prix_vente_fc = c2.number_input("Prix Vente FC", min_value=0.0)
                prix_vente_usd = c3.number_input("Prix Vente $", min_value=0.0)
                stock = c1.number_input("Stock Initial", min_value=0)
                if st.form_submit_button("💾 Ajouter Article"):
                    try:
                        data_insert = {
                            "nom_article": str(nom),
                            "categorie": str(cat),
                            "prix_achat": float(prix_achat_fc),
                            "prix_vente": float(prix_vente_fc),
                            "stock": int(stock),
                            "code_qr": str(code_qr) if code_qr else None
                        }
                        colonnes_articles = get_table_columns("articles")
                        if "prix_vente_usd" in colonnes_articles:
                            data_insert["prix_vente_usd"] = float(prix_vente_usd)
                        supabase.table("articles").insert(data_insert).execute()
                        st.success(f"Article {nom} ajouté")
                        if 'qr_code_temp' in st.session_state:
                            del st.session_state.qr_code_temp
                        st.cache_data.clear()
                        st.rerun()
                    except Exception as e:
                        st.error("Erreur ajout")
                        st.code(repr(e))

        with tab_mvt:
            st.subheader("📈 Mouvements de Stock Commerce")
            try:
                mvts = supabase.table('mouvements_stock').select("*").order("created_at", desc=True).limit(50).execute().data
            except:
                mvts = []
            
            if not mvts:
                st.info("Aucun mouvement enregistré")
            else:
                df_mvt = pd.DataFrame(mvts)
                st.dataframe(df_mvt[['article_nom', 'type', 'quantite', 'motif', 'created_by', 'created_at']], use_container_width=True, hide_index=True)

        with tab_pertes:
            st.subheader("⚠️ Déclarer Perte/Casse Article Commerce")
            
            articles_dispo = df_articles[df_articles['stock'] > 0].copy() if not df_articles.empty else pd.DataFrame()
            
            if articles_dispo.empty:
                st.warning("Aucun article en stock pour déclarer une perte")
            else:
                col1, col2 = st.columns(2)
                with col1:
                    article_dict = {f"{a['nom_article']} - Stock:{int(a['stock'])}": a for _, a in articles_dispo.iterrows()}
                    article_choisi = st.selectbox("Article abîmé/perdu", list(article_dict.keys()))
                    qte_perte = st.number_input("Quantité abîmée", min_value=1, max_value=int(article_dict[article_choisi]['stock']) if article_choisi else 1)
                with col2:
                    motif_perte = st.selectbox("Motif", ["Casse", "Vol", "Péremption", "Défaut fabrication", "Accident", "Autre"])
                    detail_perte = st.text_area("Détails", placeholder="Ex: Carton mouillé lors livraison")
                    responsable = st.text_input("Déclaré par", value=st.session_state.user_name)
                
                if article_choisi:
                    article_data = article_dict[article_choisi]
                    valeur_perte = qte_perte * float(article_data.get('prix_achat', 0))
                    st.error(f"💸 Valeur de la perte : {valeur_perte:,.0f} FC")
                
                if st.button("🚨 ENREGISTRER LA PERTE", type="primary", width="stretch"):
                    if article_choisi and qte_perte > 0:
                        article_data = article_dict[article_choisi]
                        try:
                            # 1. Déduire du stock
                            nouveau_stock = int(article_data['stock']) - qte_perte
                            supabase.table('articles').update({"stock": nouveau_stock}).eq("id", int(article_data['id'])).execute()
                            
                            # 2. Enregistrer mouvement
                            supabase.table('mouvements_stock').insert({
                                "article_id": int(article_data['id']),
                                "article_nom": str(article_data['nom_article']),
                                "type": "PERTE",
                                "quantite": -int(qte_perte),
                                "motif": f"{motif_perte} - {detail_perte}",
                                "valeur": float(valeur_perte),
                                "created_by": responsable,
                                "created_at": datetime.now().isoformat()
                            }).execute()
                            
                            st.success(f"✅ Perte enregistrée. Nouveau stock {article_data['nom_article']}: {nouveau_stock}")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error("Erreur enregistrement perte")
                            st.code(repr(e))
            
            st.divider()
            st.subheader("📋 Historique Pertes Commerce")
            try:
                pertes = supabase.table('mouvements_stock').select("*").eq("type", "PERTE").order("created_at", desc=True).limit(20).execute().data
            except:
                pertes = []
            
            if not pertes:
                st.info("Aucune perte enregistrée")
            else:
                total_pertes = sum(p.get('valeur', 0) for p in pertes)
                st.metric("💸 TOTAL PERTES COMMERCE", f"{total_pertes:,.0f} FC")
                
                for p in pertes:
                    with st.expander(f"🔴 {p.get('article_nom')} - {abs(p.get('quantite',0))} - {p.get('created_at','')[:10]}"):
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.write(f"**Qté perdue:** {abs(p.get('quantite', 0))}")
                            st.write(f"**Valeur:** {p.get('valeur', 0):,.0f} FC")
                        with col2:
                            st.write(f"**Motif:** {p.get('motif', 'N/A')}")
                            st.write(f"**Par:** {p.get('created_by', 'N/A')}")
                        with col3:
                            if st.session_state.user_role == "PDG":
                                if st.button("🗑️ Supprimer", key=f"del_perte_com_{p.get('id')}"):
                                    supabase.table('mouvements_stock').delete().eq("id", p.get('id')).execute()
                                    st.rerun()
if "🏠 Immobilier" in tab_map:
    with tab_map["🏠 Immobilier"]:
        st.markdown("## 🏠 Immobilier - Générer Facture")
        nom_client = st.text_input("👤 Nom du client", key="nom_client_bien")
        tel_client = st.text_input("Téléphone Client", value="+243...", key="tel_client_bien")
        col1, col2, col3 = st.columns(3)
        with col1:
            type_bien = st.selectbox("Type", ["Maison", "Appartement", "Bureau", "Terrain"], key="type_bien")
            adresse = st.text_input("Adresse", key="adresse_bien")
        with col2:
            prix = st.number_input("💰 Loyer USD", min_value=0.0, key="prix_bien")
            electricite = st.number_input("⚡ Électricité USD", min_value=0.0, key="elec_bien")
        with col3:
            eau = st.number_input("💧 Eau USD", min_value=0.0, key="eau_bien")
            duree_contrat = st.text_input("📅 Durée", placeholder="Ex: 6 mois", key="duree_bien")
        total_mensuel = float(prix) + float(electricite) + float(eau)
        st.info(f"💎 **TOTAL : {total_mensuel:,.2f} USD**")
        if st.button("📄 GÉNÉRER FACTURE PDF", type="primary", width="stretch", key="btn_facture_immo"):
            if nom_client and adresse:
                details_list = [
                    {"nom": f"Loyer {type_bien} | Adresse: {adresse} | Duree: {duree_contrat}", "qte": 1, "pu": prix},
                    {"nom": f"Electricite | {type_bien} - {adresse}", "qte": 1, "pu": electricite},
                    {"nom": f"Eau | {type_bien} - {adresse}", "qte": 1, "pu": eau}
                ]
                details_text = f"LOUER: {type_bien} | Adresse: {adresse} | Duree Contrat: {duree_contrat} | Loyer: {prix} $ | Electricite: {electricite} $ | Eau: {eau} $"
                periode = date.today().strftime("%B %Y")
                num_fact, pdf_bytes = creer_facture_auto("Loyer", nom_client, details_text, total_mensuel, "$", details_list, tel_client, periode, "Proforma")
                st.success(f"✅ Facture générée : {num_fact}")
                st.download_button(
                    label="📥 Télécharger Facture PDF",
                    data=bytes(pdf_bytes),
                    file_name=f"{num_fact}.pdf",
                    mime="application/pdf",
                    width="stretch",
                    key="dl_facture_immo"
                )
                pdf_b64 = base64.b64encode(pdf_bytes).decode()
                st.components.v1.html(f"""
                    <button onclick="printPDF()" style="width:100%; padding:10px; background:#00ff41; color:black; font-weight:bold; border:none; border-radius:5px; cursor:pointer; margin-top:10px;">
                        🖨️ IMPRIMER LA FACTURE
                    </button>
                    <script>
                    function printPDF() {{
                        const pdfData = 'data:application/pdf;base64,{pdf_b64}';
                        const win = window.open('', '_blank');
                        win.document.write('<iframe src="' + pdfData + '" width="100%" height="100%" style="border:none;"></iframe>');
                        win.document.close();
                        setTimeout(() => {{ win.print(); }}, 1000);
                    }}
                    </script>
                """, height=60)
                st.cache_data.clear()
            else:
                st.error("Nom client + Adresse obligatoires")

if "🚗 Automobile" in tab_map:
    with tab_map["🚗 Automobile"]:
        st.markdown("## 🚗 Automobile - Point de Vente")
        if 'panier_voiture' not in st.session_state:
            st.session_state.panier_voiture = []
        if 'vente_auto_finie' not in st.session_state:
            st.session_state.vente_auto_finie = False
        if 'pdf_auto' not in st.session_state:
            st.session_state.pdf_auto = None
        if 'num_fact_auto' not in st.session_state:
            st.session_state.num_fact_auto = None
        if 'client_auto_nom' not in st.session_state:
            st.session_state.client_auto_nom = ""
        if 'client_auto_tel' not in st.session_state:
            st.session_state.client_auto_tel = "+243..."
        if df_voitures.empty:
            st.error("Aucune voiture disponible - Ajoute des voitures dans Gestion Parc")
        else:
            col_gauche, col_droite = st.columns([2,1])
            with col_gauche:
                st.subheader("👤 Client")
                st.session_state.client_auto_nom = st.text_input("Nom Client", value=st.session_state.client_auto_nom, key="nom_client_v")
                st.session_state.client_auto_tel = st.text_input("Téléphone Client", value=st.session_state.client_auto_tel, key="tel_client_v")
                st.subheader("🔍 Choisir Voiture")
                search_qr = st.text_input("QR Code, Plaque, Marque ou Modèle", placeholder="Filtre la liste...", key="search_voiture_qr").strip()
                df_voitures_dispo = df_voitures[(df_voitures['statut'] == 'Disponible') & (df_voitures['quantite'] > 0)]
                if search_qr:
                    search_clean = search_qr.upper()
                    df_voitures_dispo = df_voitures_dispo[
                        df_voitures_dispo['code_qr'].str.contains(search_clean, case=False, na=False) |
                        df_voitures_dispo['plaque'].str.contains(search_clean, case=False, na=False) |
                        df_voitures_dispo['marque'].str.contains(search_clean, case=False, na=False) |
                        df_voitures_dispo['modele'].str.contains(search_clean, case=False, na=False)
                    ]
                if df_voitures_dispo.empty:
                    st.warning("⚠️ Aucune voiture disponible")
                else:
                    st.success(f"✅ {len(df_voitures_dispo)} véhicule(s) disponible(s)")
                    options_voitures = []
                    for _, v in df_voitures_dispo.iterrows():
                        options_voitures.append(f"{v['marque']} {v['modele']} {v.get('annee','')} | {v.get('couleur','')} | {v['plaque']} | Stock:{int(v.get('quantite',1))} | {v['prix']:,.0f}$ | ID:{v['id']}")
                    voiture_choisie = st.selectbox("Sélectionne le véhicule", options_voitures, key="select_voiture_unique")
                    if voiture_choisie:
                        id_choisi = int(voiture_choisie.split("ID:")[1])
                        v = df_voitures_dispo[df_voitures_dispo['id'] == id_choisi].iloc[0]
                        c1, c2, c3 = st.columns(3)
                        qte_max = int(v.get('quantite', 1))
                        qte = c1.number_input("Quantité", min_value=1, max_value=qte_max, value=1, key=f"qte_v_unique")
                        c2.metric("Stock dispo", qte_max)
                        c3.metric("Prix unitaire", f"{v['prix']:,.0f}$")
                        st.info(f"**{v['marque']} {v['modele']}** | Couleur: {v.get('couleur','N/A')} | Qualité: {v.get('qualite','N/A')} | QR: {v.get('code_qr','N/A')}")
                        if st.button("🛒 AJOUTER AU PANIER", type="primary", width="stretch", key="add_voiture_unique"):
                            existant = next((item for item in st.session_state.panier_voiture if item['id'] == int(v['id'])), None)
                            if existant:
                                if existant['qte'] + qte <= qte_max:
                                    existant['qte'] += qte
                                    st.success(f"Panier mis à jour: {existant['qte']}x")
                                else:
                                    st.error(f"Stock insuffisant! Max dispo: {qte_max}")
                            else:
                                st.session_state.panier_voiture.append({
                                    "id": int(v['id']),
                                    "nom": f"{v['marque']} {v['modele']} {v.get('annee','')}",
                                    "pu": float(v['prix']),
                                    "qte": int(qte),
                                    "plaque": v.get('plaque',''),
                                    "qualite": v.get('qualite',''),
                                    "code_qr": v.get('code_qr',''),
                                    "stock_max": qte_max
                                })
                                st.success("Ajouté au panier")
                            st.rerun()
            with col_droite:
                st.subheader("🛒 Panier Voiture")
                total_voiture = 0
                if st.session_state.vente_auto_finie and st.session_state.pdf_auto:
                    st.success(f"✅ Vente validée - {st.session_state.total_auto:,.0f} $")
                    st.info(f"📄 Facture: {st.session_state.num_fact_auto}")
                    if st.session_state.pdf_auto:
                        st.download_button(
                            label="📥 TÉLÉCHARGER LE PDF",
                            data=bytes(st.session_state.pdf_auto),
                            file_name=f"{st.session_state.num_fact_auto}.pdf",
                            mime="application/pdf",
                            width="stretch",
                            key="dl_facture_auto"
                        )
                    pdf_b64 = base64.b64encode(st.session_state.pdf_auto).decode()
                    st.components.v1.html(f"""
                        <button onclick="printPDFAuto()" style="width:100%; padding:10px; background:#00ff41; color:black; font-weight:bold; border:none; border-radius:5px; cursor:pointer; margin-top:10px;">
                            🖨️ IMPRIMER LA FACTURE
                        </button>
                        <script>
                        function printPDFAuto() {{
                            const pdfData = 'data:application/pdf;base64,{pdf_b64}';
                            const win = window.open('', '_blank');
                            win.document.write('<iframe src="' + pdfData + '" width="100%" height="100%" style="border:none;"></iframe>');
                            win.document.close();
                            setTimeout(() => {{ win.print(); }}, 1000);
                        }}
                        </script>
                    """, height=60)
                    if st.button("Nouvelle Vente", width="stretch", key="new_vente_auto"):
                        st.session_state.panier_voiture = []
                        st.session_state.vente_auto_finie = False
                        st.session_state.pdf_auto = None
                        st.session_state.num_fact_auto = None
                        st.session_state.client_auto_nom = ""
                        st.session_state.client_auto_tel = "+243..."
                        st.rerun()
                elif not st.session_state.panier_voiture:
                    st.info("Panier vide")
                else:
                    for idx, item in enumerate(st.session_state.panier_voiture):
                        col1, col2, col3, col4 = st.columns([3,1,1,1])
                        col1.write(f"**{item['nom']}** | {item.get('qualite','')} | {item['plaque']}")
                        col2.write(f"Qté: {item['qte']}")
                        col3.write(f"{item['pu'] * item['qte']:,.2f} $")
                        if col4.button("❌", key=f"del_v_{idx}"):
                            st.session_state.panier_voiture.pop(idx)
                            st.rerun()
                        total_voiture += item['pu'] * item['qte']
                    st.metric("💰 TOTAL VOITURE", f"{total_voiture:,.2f} $")
                    st.markdown(f"**Client:** {st.session_state.client_auto_nom}")
                    st.markdown(f"**Tel:** {st.session_state.client_auto_tel}")
                    if st.button("✅ FINALISER VENTE VOITURE", type="primary", width="stretch"):
                        if st.session_state.client_auto_nom and st.session_state.panier_voiture:
                            try:
                                details_list = [{"nom": f"{item['nom']} | {item.get('qualite','')} | {item['plaque']}",
                                                "qte": item['qte'], "pu": item['pu']} for item in st.session_state.panier_voiture]
                                details_text = " | ".join([f"{item['qte']}x {item['nom']} ({item.get('qualite','')})" for item in st.session_state.panier_voiture])
                                num_fact, pdf_bytes = creer_facture_auto("Vente Voiture", st.session_state.client_auto_nom, details_text, total_voiture, "$", details_list, st.session_state.client_auto_tel, "", "Proforma")
                                for item in st.session_state.panier_voiture:
                                    supabase.table("voitures").update({
                                        "quantite": item['stock_max'] - item['qte'],
                                        "statut": "Vendue" if item['stock_max'] - item['qte'] == 0 else "Disponible"
                                    }).eq("id", item['id']).execute()
                                st.session_state.vente_auto_finie = True
                                st.session_state.pdf_auto = pdf_bytes
                                st.session_state.num_fact_auto = num_fact
                                st.session_state.total_auto = total_voiture
                                st.session_state.panier_voiture = []
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erreur finalisation: {e}")
                        else:
                            st.error("Nom client obligatoire - Remplis à gauche")

if "🚘 Gestion Parc" in tab_map:
    with tab_map["🚘 Gestion Parc"]:
        st.markdown("## 🚘 Gestion Parc Automobile & Pertes")
        
        tab_ajout_v, tab_liste_v, tab_pertes_v = st.tabs(["➕ Ajouter Voiture", "📋 Liste Voitures", "⚠️ Pertes/Dégâts Voitures"])
        
        colonnes_voitures = get_table_columns("voitures")
        
        with tab_ajout_v:
            st.subheader("➕ Ajouter Nouvelle Voiture au Parc")
            with st.form("form_voiture_parc", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                marque = c1.text_input("Marque")
                modele = c2.text_input("Modèle")
                annee = c3.text_input("Année")
                data_insert = {"marque": str(marque), "modele": str(modele), "annee": str(annee)}
                if "plaque" in colonnes_voitures:
                    plaque = c1.text_input("Plaque")
                    data_insert["plaque"] = str(plaque)
                if "couleur" in colonnes_voitures:
                    couleur = c2.text_input("Couleur")
                    data_insert["couleur"] = str(couleur)
                if "kilometrage" in colonnes_voitures:
                    km = c3.number_input("Kilométrage", min_value=0, value=0)
                    data_insert["kilometrage"] = int(km)
                if "carburant" in colonnes_voitures:
                    carburant = c1.selectbox("Carburant", ["Essence", "Diesel", "Hybride", "Électrique"])
                    data_insert["carburant"] = str(carburant)
                if "boite" in colonnes_voitures:
                    boite = c2.selectbox("Boîte", ["Manuelle", "Automatique"])
                    data_insert["boite"] = str(boite)
                if "prix" in colonnes_voitures:
                    prix = c3.number_input("Prix Achat $", min_value=0.0, value=0.0)
                    data_insert["prix"] = float(prix)
                if "statut" in colonnes_voitures:
                    statut = c1.selectbox("Statut", ["Disponible", "En réparation", "Réservée", "Vendue"])
                    data_insert["statut"] = str(statut)
                if "quantite" in colonnes_voitures:
                    quantite = c2.number_input("Quantité en Stock", min_value=1, value=1)
                    data_insert["quantite"] = int(quantite)
                if "qualite" in colonnes_voitures:
                    qualite = c3.selectbox("Qualité", ["Neuf", "Occasion", "Reconditionné"])
                    data_insert["qualite"] = str(qualite)
                if "code_qr" in colonnes_voitures:
                    code_qr = c1.text_input("Code QR", placeholder="Scanner ou générer")
                    data_insert["code_qr"] = str(code_qr)
                if st.form_submit_button("💾 Ajouter Voiture"):
                    try:
                        supabase.table("voitures").insert(data_insert).execute()
                        st.success(f"Voiture {marque} {modele} ajoutée")
                        st.cache_data.clear()
                        st.rerun()
                    except Exception as e:
                        st.error("Erreur ajout")
                        st.code(repr(e))

        with tab_liste_v:
            st.subheader("📋 Liste des Voitures - Modifier/Supprimer")
            if df_voitures.empty:
                st.info("Aucune voiture")
            else:
                for _, row in df_voitures.iterrows():
                    with st.expander(f"{row['marque']} {row['modele']} - {row.get('plaque','')} - Stock:{row.get('quantite',0)} - {row.get('statut','')}"):
                        c1, c2, c3 = st.columns(3)
                        with c1:
                            new_marque = st.text_input("Marque", value=row['marque'], key=f"marque_v_{row['id']}")
                            new_modele = st.text_input("Modèle", value=row['modele'], key=f"modele_v_{row['id']}")
                            new_annee = st.text_input("Année", value=row.get('annee',''), key=f"annee_v_{row['id']}")
                        data_update = {"marque": str(new_marque), "modele": str(new_modele), "annee": str(new_annee)}
                        with c2:
                            if "plaque" in colonnes_voitures:
                                new_plaque = st.text_input("Plaque", value=row.get('plaque',''), key=f"plaque_v_{row['id']}")
                                data_update["plaque"] = str(new_plaque)
                            if "couleur" in colonnes_voitures:
                                new_couleur = st.text_input("Couleur", value=row.get('couleur',''), key=f"couleur_v_{row['id']}")
                                data_update["couleur"] = str(new_couleur)
                            if "kilometrage" in colonnes_voitures:
                                km_val = row.get('kilometrage', 0)
                                try:
                                    km_val = int(float(km_val)) if km_val else 0
                                except:
                                    km_val = 0
                                new_km = st.number_input("KM", value=km_val, key=f"km_v_{row['id']}")
                                data_update["kilometrage"] = int(new_km)
                        with c3:
                            if "carburant" in colonnes_voitures:
                                carburant_options = ["Essence", "Diesel", "Hybride", "Électrique"]
                                carb_val = row.get('carburant','Essence')
                                new_carb = st.selectbox("Carburant", carburant_options, index=carburant_options.index(carb_val) if carb_val in carburant_options else 0, key=f"carb_v_{row['id']}")
                                data_update["carburant"] = str(new_carb)
                            if "boite" in colonnes_voitures:
                                boite_options = ["Manuelle", "Automatique"]
                                boite_val = row.get('boite','Manuelle')
                                new_boite = st.selectbox("Boîte", boite_options, index=boite_options.index(boite_val) if boite_val in boite_options else 0, key=f"boite_v_{row['id']}")
                                data_update["boite"] = str(new_boite)
                            if "prix" in colonnes_voitures:
                                new_prix = st.number_input("Prix $", value=float(row.get('prix',0)), key=f"prix_v_{row['id']}")
                                data_update["prix"] = float(new_prix)
                            if "statut" in colonnes_voitures:
                                statut_options = ["Disponible", "En réparation", "Réservée", "Vendue"]
                                statut_val = row.get('statut','Disponible')
                                new_statut = st.selectbox("Statut", statut_options, index=statut_options.index(statut_val) if statut_val in statut_options else 0, key=f"statut_v_{row['id']}")
                                data_update["statut"] = str(new_statut)
                        if "quantite" in colonnes_voitures:
                            new_qte = st.number_input("Stock", value=int(row.get('quantite',1)), min_value=0, key=f"qte_v_{row['id']}")
                            data_update["quantite"] = int(new_qte)
                        if "qualite" in colonnes_voitures:
                            qualite_options = ["Neuf", "Occasion", "Reconditionné"]
                            qualite_val = row.get('qualite','Neuf')
                            new_qualite = st.selectbox("Qualité", qualite_options, index=qualite_options.index(qualite_val) if qualite_val in qualite_options else 0, key=f"qual_v_{row['id']}")
                            data_update["qualite"] = str(new_qualite)
                        if "code_qr" in colonnes_voitures:
                            new_code_qr = st.text_input("Code QR", value=row.get('code_qr',''), key=f"qr_v_{row['id']}")
                            data_update["code_qr"] = str(new_code_qr)
                        c1, c2 = st.columns(2)
                        if c1.button("✏️ Modifier", key=f"mod_v_parc_{row['id']}", width="stretch"):
                            try:
                                supabase.table("voitures").update(data_update).eq("id", int(row['id'])).execute()
                                st.success("Modifié")
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as e:
                                st.error("Erreur modif")
                                st.code(repr(e))
                        if st.session_state.user_role == "PDG" or perms.get('supprimer', False):
                            if c2.button("🗑️ Supprimer", key=f"del_v_parc_{row['id']}", width="stretch"):
                                try:
                                    supabase.table("voitures").delete().eq("id", int(row['id'])).execute()
                                    st.success("Supprimé")
                                    st.cache_data.clear()
                                    st.rerun()
                                except Exception as e:
                                    st.error("Erreur suppression")
                                    st.code(repr(e))

        with tab_pertes_v:
            st.subheader("⚠️ Déclarer Dégât/Perte Voiture")
            
            voitures_dispo = df_voitures[df_voitures.get('quantite', 1) > 0].copy() if not df_voitures.empty else pd.DataFrame()
            
            if voitures_dispo.empty:
                st.warning("Aucune voiture en stock pour déclarer un dégât")
            else:
                col1, col2 = st.columns(2)
                with col1:
                    voiture_dict = {f"{v['marque']} {v['modele']} - {v.get('plaque','')} - Stock:{int(v.get('quantite',1))}": v for _, v in voitures_dispo.iterrows()}
                    voiture_choisie = st.selectbox("Voiture endommagée/perdue", list(voiture_dict.keys()))
                    qte_perte_v = st.number_input("Quantité endommagée", min_value=1, max_value=int(voiture_dict[voiture_choisie].get('quantite',1)) if voiture_choisie else 1)
                with col2:
                    motif_perte_v = st.selectbox("Type de dégât", ["Accident", "Vol", "Incendie", "Panne moteur", "Dégât carrosserie", "Pneus crevés", "Autre"])
                    detail_perte_v = st.text_area("Détails du dégât", placeholder="Ex: Pare-choc avant enfoncé + phare cassé")
                    responsable_v = st.text_input("Déclaré par", value=st.session_state.user_name, key="resp_v")
                
                if voiture_choisie:
                    voiture_data = voiture_dict[voiture_choisie]
                    valeur_perte_v = qte_perte_v * float(voiture_data.get('prix', 0))
                    st.error(f"💸 Valeur de la perte : {valeur_perte_v:,.2f} $")
                
                if st.button("🚨 ENREGISTRER DÉGÂT VOITURE", type="primary", width="stretch"):
                    if voiture_choisie and qte_perte_v > 0:
                        voiture_data = voiture_dict[voiture_choisie]
                        try:
                            # 1. Déduire du stock
                            nouveau_stock_v = int(voiture_data.get('quantite',1)) - qte_perte_v
                            nouveau_statut = "En réparation" if nouveau_stock_v > 0 else "Endommagée"
                            supabase.table('voitures').update({
                                "quantite": nouveau_stock_v,
                                "statut": nouveau_statut
                            }).eq("id", int(voiture_data['id'])).execute()
                            
                            # 2. Enregistrer mouvement perte
                            supabase.table('mouvements_stock').insert({
                                "article_id": int(voiture_data['id']),
                                "article_nom": f"{voiture_data['marque']} {voiture_data['modele']} - {voiture_data.get('plaque','')}",
                                "type": "PERTE_VOITURE",
                                "quantite": -int(qte_perte_v),
                                "motif": f"{motif_perte_v} - {detail_perte_v}",
                                "valeur": float(valeur_perte_v),
                                "created_by": responsable_v,
                                "created_at": datetime.now().isoformat()
                            }).execute()
                            
                            st.success(f"✅ Dégât enregistré. Stock {voiture_data['marque']} {voiture_data['modele']}: {nouveau_stock_v}")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error("Erreur enregistrement dégât")
                            st.code(repr(e))
            
            st.divider()
            st.subheader("📋 Historique Dégâts/Pertes Voitures")
            try:
                pertes_v = supabase.table('mouvements_stock').select("*").eq("type", "PERTE_VOITURE").order("created_at", desc=True).limit(20).execute().data
            except:
                pertes_v = []
            
            if not pertes_v:
                st.info("Aucun dégât de voiture enregistré")
            else:
                total_pertes_v = sum(p.get('valeur', 0) for p in pertes_v)
                st.metric("💸 TOTAL PERTES VOITURES", f"{total_pertes_v:,.2f} $")
                
                for p in pertes_v:
                    with st.expander(f"🔴 {p.get('article_nom')} - {abs(p.get('quantite',0))} - {p.get('created_at','')[:10]}"):
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.write(f"**Qté endommagée:** {abs(p.get('quantite', 0))}")
                            st.write(f"**Valeur:** {p.get('valeur', 0):,.2f} $")
                        with col2:
                            st.write(f"**Motif:** {p.get('motif', 'N/A')}")
                            st.write(f"**Par:** {p.get('created_by', 'N/A')}")
                        with col3:
                            if st.session_state.user_role == "PDG":
                                if st.button("🗑️ Supprimer", key=f"del_perte_v_{p.get('id')}"):
                                    supabase.table('mouvements_stock').delete().eq("id", p.get('id')).execute()
                                    st.rerun()
if "💰 Comptabilité" in tab_map:
    with tab_map["💰 Comptabilité"]:
        st.markdown("## 💰 Comptabilité - Relevé par Catégorie")
        colonnes_compta = get_table_columns("comptes")
        with st.expander("➕ Ajouter Opération"):
            with st.form("form_compta", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                type_op = c1.selectbox("Type", ["Revenu", "Dépense"])
                cat = c2.text_input("Catégorie", placeholder="Ex: Loyer, Vente Auto, Carburant")
                montant = c3.number_input("Montant", min_value=0.0)
                data_insert = {"type": str(type_op), "categorie": str(cat), "montant": float(montant), "utilisateur": st.session_state.user_name}
                if "description" in colonnes_compta:
                    desc = c1.text_input("Description", placeholder="Ex: Loyer - Client Jean")
                    data_insert["description"] = str(desc)
                if "devise" in colonnes_compta:
                    devise = c2.selectbox("Devise", ["FC", "$", "€"])
                    data_insert["devise"] = str(devise)
                if "date" in colonnes_compta:
                    date_op = c3.date_input("Date", value=date.today())
                    data_insert["date"] = str(date_op)
                if st.form_submit_button("💾 Ajouter Opération"):
                    try:
                        supabase.table("compta").insert(data_insert).execute()
                        st.success("Opération ajoutée")
                        st.cache_data.clear()
                        st.rerun()
                    except Exception as e:
                        st.error("Erreur ajout")
                        st.code(repr(e))
        st.divider()
        if df_compta.empty:
            st.info("Aucune opération")
        else:
            df_compta_sorted = df_compta.sort_values('date', ascending=False)
            col_f1, col_f2, col_f3 = st.columns(3)
            date_debut = col_f1.date_input("📅 Date début", value=date.today() - timedelta(days=30), key="date_debut_compta")
            date_fin = col_f2.date_input("📅 Date fin", value=date.today(), key="date_fin_compta")
            filtre_nom = col_f3.text_input("👤 Nom Client", placeholder="Tape un nom...", key="filtre_nom_compta")

            df_filtre_compta = df_compta_sorted[(df_compta_sorted['date'] >= str(date_debut)) & (df_compta_sorted['date'] <= str(date_fin))]
            if filtre_nom:
                df_filtre_compta = df_filtre_compta[df_filtre_compta['description'].str.contains(filtre_nom, case=False, na=False)]

            col_t1, col_t2, col_t3 = st.columns(3)
            total_fc = df_filtre_compta[df_filtre_compta.get('devise','FC')=='FC']['montant'].sum()
            total_usd = df_filtre_compta[df_filtre_compta.get('devise','FC')=='$']['montant'].sum()
            total_eur = df_filtre_compta[df_filtre_compta.get('devise','FC')=='€']['montant'].sum()
            col_t1.metric("💵 Total FC", f"{total_fc:,.0f}")
            col_t2.metric("💵 Total USD", f"{total_usd:,.0f}")
            col_t3.metric("💵 Total EUR", f"{total_eur:,.0f}")
            st.divider()

            categories = df_filtre_compta.get('categorie', pd.Series(dtype=str)).dropna().unique()
            if len(categories) == 0:
                st.info("Aucune opération trouvée avec ces filtres")
            else:
                for cat in sorted(categories):
                    df_cat = df_filtre_compta[df_filtre_compta.get('categorie', '') == cat]
                    total_cat_fc = df_cat[df_cat.get('devise','FC')=='FC']['montant'].sum()
                    total_cat_usd = df_cat[df_cat.get('devise','FC')=='$']['montant'].sum()
                    total_cat_eur = df_cat[df_cat.get('devise','FC')=='€']['montant'].sum()
                    total_cat = total_cat_fc + total_cat_usd + total_cat_eur

                    with st.expander(f"📁 {cat} - {len(df_cat)} opérations - Total: {total_cat:,.0f}", expanded=False):
                        colonnes_affiche = ['date', 'type', 'description', 'montant', 'devise']
                        if 'utilisateur' in df_cat.columns:
                            colonnes_affiche.append('utilisateur')
                        st.dataframe(df_cat[colonnes_affiche], use_container_width=True, hide_index=True)

                        col_dl1, col_dl2 = st.columns(2)
                        excel_bytes_cat = generer_excel_pro(
                            df_cat,
                            f"Releve {cat} {date_debut}-{date_fin}",
                            df_cat[df_cat['type']=='Revenu']['montant'].sum(),
                            df_cat[df_cat['type']=='Dépense']['montant'].sum(),
                            df_cat[df_cat['type']=='Revenu']['montant'].sum() - df_cat[df_cat['type']=='Dépense']['montant'].sum()
                        )
                        safe_cat = str(cat).replace(" ", "_").replace("/", "_")
                        col_dl1.download_button(
                            label=f"📥 {cat} - EXCEL",
                            data=excel_bytes_cat,
                            file_name=f"Compta_{safe_cat}_{date_debut}_{date_fin}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            width="stretch",
                            key=f"dl_excel_compta_{safe_cat}_{date_debut}_{filtre_nom}"
                        )

                        pdf_cat = FPDF()
                        pdf_cat.add_page()
                        pdf_cat.set_fill_color(20, 50, 40)
                        pdf_cat.rect(0, 0, 210, 35, 'F')
                        pdf_cat.set_text_color(255, 255, 255)
                        pdf_cat.set_font("Arial", "B", 20)
                        pdf_cat.set_xy(10, 8)
                        pdf_cat.cell(0, 10, "ASYMAS BUSINESS", ln=True)
                        pdf_cat.set_font("Arial", "", 9)
                        pdf_cat.set_xy(10, 16)
                        pdf_cat.cell(0, 5, "Beni, Nord-Kivu, RDC | Tel: +243 995 105 623", ln=True)
                        pdf_cat.set_font("Arial", "B", 10)
                        pdf_cat.set_xy(150, 8)
                        filtre_txt = f"Filtre: {filtre_nom}" if filtre_nom else "Tous"
                        pdf_cat.cell(50, 6, f"Periode: {date_debut} au {date_fin}", ln=True, align="R")
                        pdf_cat.set_xy(150, 14)
                        pdf_cat.cell(50, 6, filtre_txt, ln=True, align="R")
                        pdf_cat.ln(15)
                        pdf_cat.set_text_color(0, 0, 0)
                        pdf_cat.set_fill_color(255, 204, 0)
                        pdf_cat.set_font("Arial", "B", 14)
                        pdf_cat.cell(0, 10, f"RELEVE COMPTABLE - {safe_pdf_txt(cat).upper()}", ln=True, fill=True)
                        pdf_cat.ln(5)
                        pdf_cat.set_font("Arial", "B", 11)
                        pdf_cat.cell(0, 8, f"Total FC: {total_cat_fc:,.0f} | USD: {total_cat_usd:,.0f} | EUR: {total_cat_eur:,.0f}", ln=True)
                        pdf_cat.ln(3)
                        pdf_cat.set_font("Arial", "B", 9)
                        pdf_cat.cell(20, 7, "Date", 1)
                        pdf_cat.cell(20, 7, "Type", 1)
                        pdf_cat.cell(70, 7, "Description", 1)
                        pdf_cat.cell(25, 7, "Montant", 1)
                        pdf_cat.cell(15, 7, "Dev", 1)
                        pdf_cat.cell(30, 7, "Utilisateur", 1, ln=True)
                        pdf_cat.set_font("Arial", "", 8)
                        for _, row in df_cat.iterrows():
                            try:
                                pdf_cat.cell(20, 6, safe_pdf_txt(row.get('date','')), 1)
                                pdf_cat.cell(20, 6, safe_pdf_txt(row.get('type','')), 1)
                                desc = safe_pdf_txt(row.get('description',''))[:35]
                                pdf_cat.cell(70, 6, desc, 1)
                                pdf_cat.cell(25, 6, f"{row.get('montant',0):,.0f}", 1)
                                pdf_cat.cell(15, 6, safe_pdf_txt(row.get('devise','FC')), 1)
                                pdf_cat.cell(30, 6, safe_pdf_txt(row.get('utilisateur','N/A')), 1, ln=True)
                            except:
                                continue
                        pdf_bytes_cat = bytes(pdf_cat.output(dest='S'))
                        col_dl2.download_button(
                            label=f"📥 {cat} - PDF",
                            data=pdf_bytes_cat,
                            file_name=f"Compta_{safe_cat}_{date_debut}_{date_fin}.pdf",
                            mime="application/pdf",
                            width="stretch",
                            key=f"dl_pdf_compta_{safe_cat}_{date_debut}_{filtre_nom}"
                        )
                        pdf_b64 = base64.b64encode(pdf_bytes_cat).decode()
                        st.components.v1.html(f"""
                            <button onclick="printPDF_{safe_cat}()" style="width:100%; padding:10px; background:#00ff41; color:black; font-weight:bold; border:none; border-radius:5px; cursor:pointer; margin-top:10px;">
                                🖨️ IMPRIMER LE RELEVÉ {cat}
                            </button>
                            <script>
                            function printPDF_{safe_cat}() {{
                                const pdfData = 'data:application/pdf;base64,{pdf_b64}';
                                const win = window.open('', '_blank');
                                win.document.write('<iframe src="' + pdfData + '" width="100%" height="100%" style="border:none;"></iframe>');
                                win.document.close();
                                setTimeout(() => {{ win.print(); }}, 1000);
                            }}
                            </script>
                        """, height=60)

import json
from datetime import datetime
import streamlit as st
from fpdf import FPDF
import qrcode
from io import BytesIO

# ===== FONCTION PDF =====
def generer_pdf_facture_ucad(numero, client, localisation, telephone,
                             prestations, devise, modes_paiement,
                             ing_nom="SAMY TSANGYA", ing_tel="+256766515428",
                             email="asamnesstsang636@gmail.com",
                             adresse="Beni, Nord-Kivu, RDC"):
    pdf = FPDF()
    pdf.add_page()

    # ENTÊTE ASYMAS
    pdf.set_y(0)
    pdf.set_fill_color(20, 50, 40)
    pdf.rect(0, 0, 210, 32, 'F')
    pdf.set_xy(10, 2)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", 'B', 18)
    pdf.cell(130, 10, "ASYMAS CONSULTING", 0, 0, '', True)
    pdf.set_font("Arial", 'B', 11)
    pdf.cell(0, 10, "FACTURE N", 0, 1, 'R', True)
    pdf.set_x(10)
    pdf.set_font("Arial", size=10)
    pdf.cell(130, 6, f"{adresse} | Tel: {ing_tel}", 0, 0, '', True)
    pdf.cell(0, 6, f"{numero}", 0, 1, 'R', True)
    pdf.set_x(10)
    pdf.cell(130, 6, f"Email: {email}", 0, 0, '', True)
    pdf.cell(0, 6, f"Date: {datetime.now().strftime('%d/%m/%Y')}", 0, 1, 'R', True)
    pdf.set_x(10)
    pdf.set_font("Arial", size=9)
    pdf.cell(0, 6, "Etudes - Fournitures - Travaux Industriels Electriques & Batiment", ln=True, align='C', fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(5)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(8)

    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 8, "FACTURE D'HONORAIRES", ln=True, align='C')
    pdf.ln(3)

    pdf.set_font("Arial", 'B', 11)
    pdf.cell(0, 7, f"LOCALISATION: {localisation}", ln=True)
    pdf.cell(0, 7, f"CLIENT: {client}", ln=True)
    if telephone and telephone!= "+243...":
        pdf.cell(0, 7, f"TEL: {telephone}", ln=True)
    pdf.ln(5)

    total = sum(float(p['montant']) for p in prestations)

    # TABLEAU DETAIL DES ECHEANCES
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(0, 6, "Détail des échéances", ln=True, align='C')
    pdf.ln(2)

    pdf.set_font("Arial", 'B', 8)
    pdf.set_fill_color(200, 200, 200)
    pdf.cell(8, 6, "N°", 1, 0, 'C', True)
    pdf.cell(45, 6, "Transaction", 1, 0, 'C', True)
    pdf.cell(22, 6, "Montant", 1, 0, 'C', True)
    pdf.cell(22, 6, "Date échéance", 1, 0, 'C', True)
    pdf.cell(32, 6, "Mode Règlement", 1, 0, 'C', True)
    pdf.cell(22, 6, "Date paiement", 1, 0, 'C', True)
    pdf.cell(39, 6, "N.Quittance", 1, 1, 'C', True)

    pdf.set_font("Arial", size=8)
    pdf.set_fill_color(255, 255, 255)
    for idx, item in enumerate(prestations, 1):
        montant = float(item['montant'])
        mode = item.get('mode_paiement', 'Espèces')
        pdf.cell(8, 6, str(idx), 1, 0, 'C')
        pdf.cell(45, 6, item['designation'][:23], 1)
        pdf.cell(22, 6, f"{montant:,.0f}", 1, 0, 'R')
        pdf.cell(22, 6, datetime.now().strftime('%d/%m/%Y'), 1, 0, 'C')
        pdf.cell(32, 6, mode[:16], 1, 0, 'C')
        pdf.cell(22, 6, "Délai raisonnable", 1, 0, 'C')
        pdf.cell(39, 6, f"Q-{numero[-6:]}{idx}", 1, 1, 'C')

    pdf.set_font("Arial", 'B', 8)
    pdf.cell(75, 6, "Total", 1, 0, 'R', True)
    pdf.cell(22, 6, f"{total:,.0f}", 1, 0, 'R', True)
    pdf.cell(93, 6, "", 1, 1)

    # MODES DE PAIEMENT SAISIS PAR TOI
    pdf.ln(5)
    pdf.set_font("Arial", 'I', 8)
    pdf.cell(0, 5, "Modes de paiement acceptés:", ln=True)
    pdf.set_font("Arial", size=8)
    pdf.cell(60, 5, "Mode de règlement", 1, 0, 'C')
    pdf.cell(130, 5, "Détails / N° Compte", 1, 1, 'C')
    for mode_nom, mode_detail in modes_paiement:
        pdf.cell(60, 5, mode_nom, 1, 0, 'C')
        pdf.cell(130, 5, mode_detail[:58], 1, 1, 'C')

    pdf.ln(5)
    pdf.set_font("Arial", 'B', 11)
    pdf.set_fill_color(255, 200, 0)
    pdf.cell(100, 8, f"TOTAL GENERAL ({devise})", 1, 0, 'R', True)
    pdf.cell(90, 8, f"{total:,.2f}", 1, 1, 'R', True)

    # QR CODE
    qr_data = {
        "numero": numero,
        "client": client,
        "total": total,
        "devise": devise,
        "prestations": [{"des": p['designation'], "montant": p['montant'], "mode": p.get('mode_paiement')} for p in prestations],
        "modes_paiement": modes_paiement
    }
    qr = qrcode.QRCode(box_size=3, border=1)
    qr.add_data(json.dumps(qr_data))
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    y_pos = pdf.get_y() + 10
    pdf.image(buf, x=160, y=y_pos, w=35)
    pdf.set_xy(160, y_pos + 36)
    pdf.set_font("Arial", 'I', 7)
    pdf.cell(35, 4, "Scan pour vérifier", align='C')

    pdf.set_xy(15, y_pos)
    pdf.set_font("Arial", 'B', 11)
    pdf.cell(0, 8, "SIGNATURE INGENIEUR RESPONSABLE:", ln=True)
    pdf.ln(15)
    pdf.line(15, pdf.get_y(), 105, pdf.get_y())
    pdf.ln(5)
    pdf.set_font("Arial", size=10)
    pdf.cell(0, 6, f"Ing. {ing_nom}", ln=True)
    pdf.cell(0, 6, f"Tel: {ing_tel}", ln=True)

    out = pdf.output(dest='S')
    return bytes(out), total

# ===== BLOC STREAMLIT =====
if "📄 Factures" in tab_map:
    with tab_map["📄 Factures"]:
        st.markdown("## 📄 Factures - Honoraires")

        if st.button("➕ Nouvelle Facture", type="primary"):
            st.session_state.nouvelle_facture = True

        if st.session_state.get("nouvelle_facture", False):
            st.subheader("Créer une Facture d'Honoraires")

            col1, col2 = st.columns(2)
            with col1:
                client_fac = st.text_input("Client", key="client_fac")
                localisation_fac = st.text_input("Localisation", value="Beni, RDC", key="loc_fac")
            with col2:
                tel_fac = st.text_input("Téléphone", placeholder="+243...", key="tel_fac")
                devise_fac = st.selectbox("Devise", ["USD", "FC", "€"], key="devise_fac")

            # SAISIE DES 3 OU 4 MODES DE PAIEMENT AVEC LEURS INFOS
            st.markdown("**Modes de paiement acceptés - saisissez 3 à 4 modes avec leurs détails**")
            modes_paiement = []
            for i in range(4):
                col_m1, col_m2 = st.columns([2, 3])
                with col_m1:
                    mode_nom = st.text_input(f"Mode {i+1}", key=f"mode_nom_{i}", placeholder="Ex: Espèces")
                with col_m2:
                    mode_detail = st.text_input(f"Détail {i+1}", key=f"mode_detail_{i}", placeholder="Ex: 0812... ou N° Compte BIA")
                if mode_nom and mode_detail:
                    modes_paiement.append((mode_nom, mode_detail))

            # Liste des noms de modes pour les selectbox
            choix_mode = [m[0] for m in modes_paiement] if modes_paiement else ["Espèces"]

            st.markdown("**Prestations réalisées**")

            choix_designation = [
                "Frais d'étude technique",
                "Main d'oeuvre",
                "Visite de site",
                "Rapport technique",
                "Déplacement",
                "Autre"
            ]

            prestations = []
            for i in range(4):
                st.markdown(f"**Désignation {i+1}**")
                col_a, col_b, col_c, col_d = st.columns([2.5, 2.5, 1.5, 2])

                with col_a:
                    des_choix = st.selectbox(
                        "Désignation",
                        choix_designation,
                        key=f"des_choix_fac_{i}",
                        label_visibility="collapsed"
                    )
                    if des_choix == "Autre":
                        des = st.text_input("Saisir", key=f"des_custom_fac_{i}", label_visibility="collapsed")
                    else:
                        des = des_choix

                with col_b:
                    detail = st.text_input("Détail", key=f"detail_fac_{i}", label_visibility="collapsed")

                with col_c:
                    montant = st.number_input("Montant", min_value=0.0, value=0.0, step=1000.0, key=f"montant_fac_{i}", label_visibility="collapsed")

                with col_d:
                    mode = st.selectbox("Mode paiement", choix_mode, key=f"mode_fac_{i}", label_visibility="collapsed")

                if des and montant > 0:
                    prestations.append({
                        "designation": des,
                        "detail": detail,
                        "montant": montant,
                        "mode_paiement": mode
                    })

            col_gen, col_ann = st.columns(2)
            with col_gen:
                if st.button("📄 Générer Facture PDF", type="primary", width="stretch"):
                    if client_fac and prestations and modes_paiement:
                        numero_fac = f"FAC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                        pdf_bytes, total = generer_pdf_facture_ucad(
                            numero_fac, client_fac, localisation_fac, tel_fac,
                            prestations, devise_fac, modes_paiement
                        )
                        st.download_button(
                            "⬇️ Télécharger Facture PDF",
                            data=pdf_bytes,
                            file_name=f"{numero_fac}.pdf",
                            mime="application/pdf",
                            width="stretch"
                        )
                        st.success(f"Facture {numero_fac} - Total: {total:,.2f} {devise_fac}")
                    else:
                        st.error("Client, au moins 1 prestation, et au moins 1 mode de paiement requis")
            with col_ann:
                if st.button("❌ Annuler", width="stretch"):
                    st.session_state.nouvelle_facture = False
                    st.rerun()

            st.divider()
import json
from datetime import datetime
from fpdf import FPDF
import streamlit as st
import qrcode
from io import BytesIO

# ===== FONCTION GENERATION PDF =====
def generer_pdf_devis_consulting(numero, type_devis, client, titre, parcelle, localisation,
                                 sections, devise, telephone, main_oeuvre,
                                 ing_nom="SAMY TSANGYA", ing_tel="+256766515428",
                                 email="asamnesstsang636@gmail.com", adresse="Beni, Nord-Kivu, RDC"):
    pdf = FPDF()
    pdf.add_page()

    # BANDEAU VERT FONCE collé au bord haut
    pdf.set_y(0)
    pdf.set_fill_color(20, 50, 40)
    pdf.rect(0, 0, 210, 32, 'F')

    pdf.set_xy(10, 2)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", 'B', 18)
    pdf.cell(130, 10, "ASYMAS CONSULTING", 0, 0, '', True)
    pdf.set_font("Arial", 'B', 11)
    pdf.cell(0, 10, "DEVIS N", 0, 1, 'R', True)

    pdf.set_x(10)
    pdf.set_font("Arial", size=10)
    pdf.cell(130, 6, f"{adresse} | Tel: {ing_tel}", 0, 0, '', True)
    pdf.cell(0, 6, f"{numero}", 0, 1, 'R', True)

    pdf.set_x(10)
    pdf.cell(130, 6, f"Email: {email}", 0, 0, '', True)
    pdf.cell(0, 6, f"Date: {datetime.now().strftime('%d/%m/%Y')}", 0, 1, 'R', True)

    pdf.set_x(10)
    pdf.set_font("Arial", size=9)
    pdf.cell(0, 6, "Etudes - Fournitures - Travaux Industriels Electriques & Batiment", ln=True, align='C', fill=True)

    pdf.set_text_color(0, 0, 0)
    pdf.ln(5)
    pdf.set_draw_color(0, 0, 0)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(8)

    # Titre centré
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 8, titre.upper(), ln=True, align='C')
    pdf.ln(3)

    # Infos client
    pdf.set_font("Arial", 'B', 11)
    pdf.cell(0, 7, f"LOCALISATION: {localisation}", ln=True)
    pdf.cell(0, 7, f"CLIENT: {client}", ln=True)
    if telephone and telephone!= "+243...":
        pdf.cell(0, 7, f"TEL: {telephone}", ln=True)
    pdf.ln(5)

    total_general = 0
    for section in sections:
        if pdf.get_y() > 240:
            pdf.add_page()

        # EN-TETE TABLEAU avec colonne DETAIL
        pdf.set_font("Arial", 'B', 9)
        pdf.set_fill_color(200, 200, 200)
        pdf.cell(8, 7, "N", 1, 0, 'C', True)
        pdf.cell(55, 7, "DESIGNATION", 1, 0, 'C', True)
        pdf.cell(45, 7, "DETAIL", 1, 0, 'C', True)
        pdf.cell(15, 7, "Unité", 1, 0, 'C', True)
        pdf.cell(18, 7, "Qté", 1, 0, 'C', True)
        pdf.cell(22, 7, "Prix U", 1, 0, 'C', True)
        pdf.cell(27, 7, "Prix total", 1, 1, 'C', True)

        # Ligne section
        pdf.set_font("Arial", 'B', 10)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(8, 6, section.get('numero', 'A'), 1, 0, 'C', True)
        pdf.cell(182, 6, section.get('titre', 'INDUSTRIAL'), 1, 1, 'L', True)

        # Articles fond blanc
        pdf.set_font("Arial", size=9)
        pdf.set_fill_color(255, 255, 255)
        sous_total_sec = 0
        for item in section.get('items', []):
            if pdf.get_y() > 265:
                pdf.add_page()
            pt = float(item.get('qte', 0)) * float(item.get('pu', 0))
            sous_total_sec += pt
            pdf.cell(8, 6, str(item.get('num', '')), 1)
            pdf.cell(55, 6, item.get('designation', '')[:28], 1)
            pdf.cell(45, 6, item.get('detail', '')[:22], 1)
            pdf.cell(15, 6, str(item.get('unite', '')), 1, 0, 'C')
            pdf.cell(18, 6, f"{float(item.get('qte', 0)):,.2f}", 1, 0, 'R')
            pdf.cell(22, 6, f"{float(item.get('pu', 0)):,.2f}", 1, 0, 'R')
            pdf.cell(27, 6, f"{pt:,.2f}", 1, 1, 'R')

        # Sous Total
        pdf.set_font("Arial", 'B', 10)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(163, 7, "Sous Total", 1, 0, 'R', True)
        pdf.cell(27, 7, f"{sous_total_sec:,.2f}", 1, 1, 'R', True)
        pdf.ln(1)
        total_general += sous_total_sec

    # Main d'oeuvre si > 0
    if main_oeuvre > 0:
        pdf.set_font("Arial", 'B', 10)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(163, 7, "Main d'oeuvre", 1, 0, 'R', True)
        pdf.cell(27, 7, f"{main_oeuvre:,.2f}", 1, 1, 'R', True)

    # TOTAL GENERAL JAUNE
    if pdf.get_y() > 230:
        pdf.add_page()
    pdf.set_font("Arial", 'B', 11)
    pdf.set_fill_color(255, 200, 0)
    pdf.cell(163, 9, f"TOTAL GENERAL ({devise})", 1, 0, 'R', True)
    pdf.cell(27, 9, f"{total_general + main_oeuvre:,.2f}", 1, 1, 'R', True)

    # ===== QR CODE =====
    # Données encodées dans le QR
    qr_data = {
        "numero": numero,
        "client": client,
        "titre": titre,
        "total": total_general + main_oeuvre,
        "devise": devise,
        "date": datetime.now().strftime('%d/%m/%Y'),
        "verif": f"https://asymas.com/verifier/{numero}" # change par ton URL
    }
    qr = qrcode.QRCode(version=1, box_size=3, border=1)
    qr.add_data(json.dumps(qr_data))
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    # Insérer QR en mémoire
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)

    y_pos = pdf.get_y() + 10
    if y_pos > 220: # nouvelle page si pas de place
        pdf.add_page()
        y_pos = 20

    pdf.image(buf, x=160, y=y_pos, w=35) # QR en bas à droite
    pdf.set_xy(160, y_pos + 36)
    pdf.set_font("Arial", 'I', 7)
    pdf.cell(35, 4, "Scan pour vérifier", align='C')

    # Signature à gauche
    pdf.set_xy(15, y_pos)
    pdf.set_font("Arial", 'B', 11)
    pdf.cell(0, 8, "SIGNATURE INGENIEUR RESPONSABLE:", ln=True)
    pdf.ln(15)
    pdf.line(15, pdf.get_y(), 105, pdf.get_y())
    pdf.ln(5)
    pdf.set_font("Arial", size=10)
    pdf.cell(0, 6, f"Ing. {ing_nom}", ln=True)
    pdf.cell(0, 6, f"Tel: {ing_tel}", ln=True)
    pdf.cell(0, 6, f"Adresse: {adresse}", ln=True)
    pdf.ln(8)
    pdf.set_text_color(0, 150, 0)
    pdf.set_font("Arial", 'I', 9)
    pdf.cell(0, 6, "Devis estimatif - Valable 30 jours", ln=True, align='C')

    out = pdf.output(dest='S')
    return bytes(out)

import streamlit as st
import json
import base64
from datetime import datetime

# ===== FONCTIONS PDF =====
def generer_pdf_facture_consulting(numero, client, titre, date, ref_devis, sections, devise, total, retenue, net, ing_nom, ing_tel):
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "FACTURE A HONORER - TRAVAUX EXECUTES", 0, 1, "C")
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 6, f"N°: {numero} | Date: {date} | Ref Devis: {ref_devis}", 0, 1, "C")
    pdf.ln(5)
    pdf.cell(0, 6, f"Client: {client}", 0, 1)
    pdf.cell(0, 6, f"Ingénieur: {ing_nom} | Tel: {ing_tel}", 0, 1)
    pdf.ln(5)

    # Tableau
    pdf.set_font("Arial", "B", 8)
    pdf.cell(10, 6, "N°", 1)
    pdf.cell(70, 6, "DESIGNATION", 1)
    pdf.cell(15, 6, "UNITE", 1)
    pdf.cell(20, 6, "QTE TOT", 1)
    pdf.cell(20, 6, "QTE EXE", 1)
    pdf.cell(20, 6, "PU", 1)
    pdf.cell(25, 6, "MONTANT", 1, 1)

    pdf.set_font("Arial", "", 8)
    for section in sections:
        pdf.cell(0, 6, f"{section.get('numero','')}. {section.get('titre','')}", 1, 1)
        for item in section.get('items',[]):
            montant = item.get('qte_execute',0) * item.get('pu',0)
            pdf.cell(10, 6, str(item.get('num','')), 1)
            pdf.cell(70, 6, item.get('designation',''), 1)
            pdf.cell(15, 6, item.get('unite',''), 1)
            pdf.cell(20, 6, str(item.get('qte_totale',0)), 1)
            pdf.cell(20, 6, str(item.get('qte_execute',0)), 1)
            pdf.cell(20, 6, f"{item.get('pu',0):.2f}", 1)
            pdf.cell(25, 6, f"{montant:.2f}", 1, 1)

    pdf.ln(3)
    pdf.cell(135, 6, "TOTAL TRAVAUX EXECUTES:", 1)
    pdf.cell(45, 6, f"{total:,.2f} {devise}", 1, 1, "R")
    pdf.cell(135, 6, "RETENUE GARANTIE:", 1)
    pdf.cell(45, 6, f"{retenue:,.2f} {devise}", 1, 1, "R")
    pdf.cell(135, 6, "NET A PAYER:", 1)
    pdf.cell(45, 6, f"{net:,.2f} {devise}", 1, 1, "R")

    return bytes(pdf.output())

# ===== INTERFACE STREAMLIT =====
if "📋 Devis" in tab_map:
    with tab_map["📋 Devis"]:
        st.markdown("## 📋 Devis Consulting - Industriel & Bâtiment")
        perms = st.session_state.user_perms
        is_pdg = st.session_state.user_role == "PDG"
        from datetime import datetime
        import json
        import base64

        # === INIT SESSION STATE ===
        if 'devis_sections_ind' not in st.session_state: st.session_state.devis_sections_ind = []
        if 'devis_bat_sections' not in st.session_state: st.session_state.devis_bat_sections = []
        if 'devis_bat_titre' not in st.session_state: st.session_state.devis_bat_titre = "DEVIS BATIMENT"
        if 'devis_bat_main_oeuvre' not in st.session_state: st.session_state.devis_bat_main_oeuvre = 0.0
        if 'facture_bat_sections' not in st.session_state: st.session_state.facture_bat_sections = []
        if 'facture_bat_pourcentage' not in st.session_state: st.session_state.facture_bat_pourcentage = 30.0
        if 'titre_fact_bat' not in st.session_state: st.session_state.titre_fact_bat = "FACTURE A HONORER - TRAVAUX EXECUTES"

        # === VERIFIER PERMISSIONS ===
        peut_voir_ind = is_pdg or perms.get('devis_industriel', False) or perms.get('devis_industriel_download', False) or perms.get('devis_industriel_print', False) or perms.get('devis_historique_industriel', False)
        peut_voir_bat = is_pdg or perms.get('devis_batiment', False) or perms.get('devis_batiment_download', False) or perms.get('devis_batiment_print', False) or perms.get('devis_historique_batiment', False)
        peut_facture_bat = peut_voir_bat

        peut_creer_ind = is_pdg or perms.get('devis_industriel', False)
        peut_creer_bat = is_pdg or perms.get('devis_batiment', False)
        peut_dl_ind = is_pdg or perms.get('devis_industriel_download', False)
        peut_dl_bat = is_pdg or perms.get('devis_batiment_download', False)
        peut_pr_ind = is_pdg or perms.get('devis_industriel_print', False)
        peut_pr_bat = is_pdg or perms.get('devis_batiment_print', False)
        peut_hist_ind = is_pdg or perms.get('devis_historique_industriel', False)
        peut_hist_bat = is_pdg or perms.get('devis_historique_batiment', False)

        if not peut_voir_ind and not peut_voir_bat:
            st.warning("🔒 Vous n'avez aucune permission pour les devis")
            st.stop()

        # === ONGLETS ===
        tabs_list = []
        if peut_voir_ind: tabs_list.append("🏭 Devis Industriel")
        if peut_voir_bat: tabs_list.append("🏗️ Devis Bâtiment")
        if peut_facture_bat: tabs_list.append("🧾 Facture Bâtiment")
        if peut_hist_ind or peut_hist_bat: tabs_list.append("📜 Historique")

        tabs = st.tabs(tabs_list)
        tab_idx = 0

        # ===== 1. ONGLET INDUSTRIEL =====
        if peut_voir_ind:
            with tabs[tab_idx]:
                if peut_creer_ind:
                    st.subheader("🏭 Nouveau Devis Industriel")
                    st.session_state.devis_type = "Industriel"
                    col_ing1, col_ing2 = st.columns(2)
                    with col_ing1: ing_nom_ind = st.text_input("👨‍🔧 Ingénieur", value="SAMY TSANGYA", key="ing_nom_ind")
                    with col_ing2: ing_tel_ind = st.text_input("📞 Tél Ingénieur", value="+256766515428", key="ing_tel_ind")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        client_devis_ind = st.text_input("👤 Client", value="EMPIRE TECHNOLOGY", key="client_devis_ind")
                        tel_client_devis_ind = st.text_input("📞 Téléphone", value="+243971409712", key="tel_devis_ind")
                    with col2:
                        titre_devis_ind = st.text_input("📋 Titre Projet", value="TRAITEMENT ET PRODUCTION EAUX MINERALE", key="titre_devis_ind")
                        parcelle_devis_ind = st.text_input("🗺️ Parcelle N°", key="parcelle_devis_ind")
                    with col3:
                        localisation_devis_ind = st.text_input("📍 Localisation", value="BENI-DRCONGO", key="loc_devis_ind")
                        devise_devis_ind = st.selectbox("💵 Devise", ["USD", "FC", "€"], key="devise_devis_ind")
                    st.divider()
                    if not st.session_state.devis_sections_ind:
                        items = [{"num": "1", "designation": "OZONEUR", "detail": "", "unite": "pc", "qte": 1.00, "pu": 450},{"num": "2", "designation": "LAMPE UV", "detail": "", "unite": "pc", "qte": 2.00, "pu": 200},{"num": "3", "designation": "DECAPELLE", "detail": "", "unite": "m", "qte": 3.00, "pu": 120}] + [{"num": str(i+4), "designation": "", "detail": "", "unite": "pc", "qte": 0, "pu": 0} for i in range(7)]
                        st.session_state.devis_sections_ind = [{"numero": "A", "titre": "INDUSTRIAL", "items": items}]
                    total_general_ind = 0
                    for idx, section in enumerate(st.session_state.devis_sections_ind):
                        col_titre1, col_titre2 = st.columns([0.2, 3])
                        with col_titre1: section['numero'] = st.text_input("N°Sec", value=section['numero'], key=f"numsec_ind_{idx}", label_visibility="collapsed")
                        with col_titre2: section['titre'] = st.text_input("Titre Section", value=section['titre'], key=f"titresec_ind_{idx}", label_visibility="collapsed")
                        st.markdown(f"**{section['numero']}. {section['titre']}**")
                        sous_total_sec = 0
                        items_to_delete = []
                        for i, item in enumerate(section['items']):
                            col1, col2, col3, col4, col5, col6, col7, col8 = st.columns([0.5, 2.5, 2, 1, 1, 1, 1, 0.5])
                            with col1: section['items'][i]['num'] = st.text_input("N°", value=item.get('num', ''), key=f"num_ind_{idx}_{i}", label_visibility="collapsed")
                            with col2: section['items'][i]['designation'] = st.text_input("Désignation", value=item.get('designation', ''), key=f"des_ind_{idx}_{i}", label_visibility="collapsed")
                            with col3: section['items'][i]['detail'] = st.text_input("Détail", value=item.get('detail', ''), key=f"detail_ind_{idx}_{i}", label_visibility="collapsed")
                            with col4: section['items'][i]['unite'] = st.selectbox("Unité", ["m", "pc", "kg", "lot"], index=["m", "pc", "kg", "lot"].index(item.get('unite', 'pc')), key=f"unit_ind_{idx}_{i}", label_visibility="collapsed")
                            with col5: section['items'][i]['qte'] = st.number_input("Qté", value=float(item.get('qte', 0)), min_value=0.0, key=f"qte_ind_{idx}_{i}", label_visibility="collapsed", format="%.2f")
                            with col6: section['items'][i]['pu'] = st.number_input("PU", value=float(item.get('pu', 0)), min_value=0.0, key=f"pu_ind_{idx}_{i}", label_visibility="collapsed", format="%.2f")
                            with col7:
                                pt = section['items'][i]['qte'] * section['items'][i]['pu']; st.markdown(f"**{pt:,.2f}**"); sous_total_sec += pt
                            with col8:
                                if st.button("❌", key=f"del_item_ind_{idx}_{i}"): items_to_delete.append(i)
                        for i in sorted(items_to_delete, reverse=True): section['items'].pop(i); st.rerun()
                        total_general_ind += sous_total_sec; st.markdown(f"**Sous-total: {sous_total_sec:,.2f}**"); st.divider()
                    main_oeuvre_ind = st.number_input("👷 Main d'oeuvre", min_value=0.0, value=0.0, key="mo_devis_ind")
                    cout_total_ind = total_general_ind + main_oeuvre_ind; st.metric("COUT TOTAL DU PROJET", f"{cout_total_ind:,.2f} {devise_devis_ind}")
                    if st.button("📄 GÉNÉRER DEVIS PDF", type="primary", width="stretch", key="gen_devis_ind"):
                        if client_devis_ind and titre_devis_ind:
                            numero_devis = f"DEV-IND-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                            try:
                                data_devis = {"numero": numero_devis, "type": "Industriel", "client": client_devis_ind, "telephone": tel_client_devis_ind, "titre": titre_devis_ind, "parcelle": parcelle_devis_ind, "localisation": localisation_devis_ind, "sections": st.session_state.devis_sections_ind, "main_oeuvre": main_oeuvre_ind, "total": cout_total_ind, "devise": devise_devis_ind, "created_by": st.session_state.user_name, "created_at": datetime.now().isoformat()}
                                supabase.table('devis').insert(data_devis).execute(); st.success(f"✅ Devis enregistré : {numero_devis}")
                            except Exception as e: st.error("Erreur enregistrement"); st.exception(e); st.stop()
                            pdf_bytes = generer_pdf_devis_consulting(numero_devis, "Industriel", client_devis_ind, titre_devis_ind, parcelle_devis_ind, localisation_devis_ind, st.session_state.devis_sections_ind, devise_devis_ind, tel_client_devis_ind, main_oeuvre_ind, ing_nom_ind, ing_tel_ind)
                            st.session_state.pdf_devis_ind = pdf_bytes; st.session_state.num_devis_ind = numero_devis; st.rerun()
                        else: st.error("Client et Titre requis")
                else: st.info("🔒 Vous n'avez pas la permission de créer des devis industriels")
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    if 'pdf_devis_ind' in st.session_state and st.session_state.pdf_devis_ind and peut_dl_ind:
                        st.download_button("📥 Télécharger PDF", data=st.session_state.pdf_devis_ind, file_name=f"{st.session_state.num_devis_ind}.pdf", mime="application/pdf", key="dl_devis_ind_1", width="stretch")
                with col_btn2:
                    if st.button("🔄 Nouveau devis Industriel", key="reset_devis_ind_1"): st.session_state.devis_sections_ind = []
                    if 'pdf_devis_ind' in st.session_state: del st.session_state.pdf_devis_ind; st.rerun()
            tab_idx += 1

        # ===== 2. ONGLET BATIMENT =====
        if peut_voir_bat:
            with tabs[tab_idx]:
                if peut_creer_bat:
                    st.subheader("🏗️ Nouveau Devis Bâtiment - ASYMAS CONSULTING")
                    ing_nom_bat = "ESDRAS"; ing_tel_bat = "+243 972 888 690"
                    st.info(f"**Ingénieur Bâtiment:** {ing_nom_bat} | **Tél:** {ing_tel_bat} | **Email:** esdrastsangya@gmail.com")
                    if not st.session_state.devis_bat_sections:
                        st.session_state.devis_bat_sections = [{"numero": "I", "titre": "Installation chantier / Demolitions", "items": [{"num": "", "designation": "Installation chantier", "unite": "ff", "qte": 1, "pu": 200}, {"num": "", "designation": "Demolitions", "unite": "ff", "qte": 1, "pu": 70}]},{"numero": "II", "titre": "fondation", "items": [{"num": "1", "designation": "moellon", "unite": "Canters", "qte": 9, "pu": 50}, {"num": "2", "designation": "sable", "unite": "Canters", "qte": 4, "pu": 40}, {"num": "3", "designation": "ciment", "unite": "sac", "qte": 23, "pu": 13.5}]},{"numero": "III", "titre": "Élévation de mur et corniche", "items": [{"num": "1", "designation": "bloc ciment", "unite": "pièce", "qte": 987, "pu": 1}]},{"numero": "IV", "titre": "Coffrage Colonne, Cornice et Socle", "items": [{"num": "1", "designation": "socle et longrine", "unite": "pièce", "qte": 8, "pu": 7}]},{"numero": "V", "titre": "Finissage", "items": [{"num": "", "designation": "ciment", "unite": "sac", "qte": 20, "pu": 13.5}]}]
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        client_devis_bat = st.text_input("👤 Client", key="client_devis_bat")
                        tel_client_devis_bat = st.text_input("📞 Téléphone Client", value="+243...", key="tel_devis_bat")
                    with col2:
                        st.session_state.devis_bat_titre = st.text_input("📋 Titre du Devis", value=st.session_state.devis_bat_titre, key="titre_devis_bat")
                        parcelle_devis_bat = st.text_input("🗺️ Parcelle N°", key="parcelle_devis_bat")
                    with col3:
                        localisation_devis_bat = st.text_input("📍 Localisation", key="loc_devis_bat")
                        devise_devis_bat = st.selectbox("💵 Devise", ["USD", "FC", "€"], key="devise_devis_bat")
                    st.divider()
                    total_general_bat = 0
                    for idx, section in enumerate(st.session_state.devis_bat_sections):
                        col_titre1, col_titre2 = st.columns([0.2, 3])
                        with col_titre1: section['numero'] = st.text_input("N°Sec", value=section['numero'], key=f"numsec_bat_{idx}", label_visibility="collapsed")
                        with col_titre2: section['titre'] = st.text_input("Titre Section", value=section['titre'], key=f"titresec_bat_{idx}", label_visibility="collapsed")
                        st.markdown(f"**{section['numero']}. {section['titre']}**")
                        sous_total_sec = 0
                        items_to_delete = []
                        for i, item in enumerate(section.get('items', [])):
                            col1, col2, col3, col4, col5, col6, col7 = st.columns([0.5, 4, 1.5, 1.5, 1.5, 1.5, 0.5])
                            with col1: section['items'][i]['num'] = st.text_input("N°", value=str(item.get('num','')), key=f"num_bat_{idx}_{i}", label_visibility="collapsed")
                            with col2: section['items'][i]['designation'] = st.text_input("Désignation", value=item.get('designation',''), key=f"des_bat_{idx}_{i}", label_visibility="collapsed")
                            with col3: section['items'][i]['unite'] = st.selectbox("Unité", ["Canters", "sac", "pièce", "kg", "ff"], index=["Canters", "sac", "pièce", "kg", "ff"].index(item.get('unite','ff')) if item.get('unite','ff') in ["Canters", "sac", "pièce", "kg", "ff"] else 4, key=f"unit_bat_{idx}_{i}", label_visibility="collapsed")
                            with col4: section['items'][i]['qte'] = st.number_input("Qté", value=float(item.get('qte',0)), min_value=0.0, key=f"qte_bat_{idx}_{i}", label_visibility="collapsed", format="%.2f")
                            with col5: section['items'][i]['pu'] = st.number_input("PU", value=float(item.get('pu',0)), min_value=0.0, key=f"pu_bat_{idx}_{i}", label_visibility="collapsed", format="%.2f")
                            with col6: pt = section['items'][i]['qte'] * section['items'][i]['pu']; st.markdown(f"**{pt:,.2f}**"); sous_total_sec += pt
                            with col7:
                                if st.button("❌", key=f"del_item_bat_{idx}_{i}"): items_to_delete.append(i)
                        for i in sorted(items_to_delete, reverse=True): section['items'].pop(i); st.rerun()
                        total_general_bat += sous_total_sec; st.markdown(f"**Sous-total: {sous_total_sec:,.2f}**")
                        if st.button("➕ Ajouter Ligne", key=f"add_line_bat_{idx}"): section['items'].append({"num": "", "designation": "", "unite": "ff", "qte": 0, "pu": 0}); st.rerun()
                        st.divider()
                    if st.button("➕ Ajouter Section", key="add_section_bat", width="stretch"):
                        new_num = f"Sec{len(st.session_state.devis_bat_sections)+1}"
                        st.session_state.devis_bat_sections.append({"numero": new_num, "titre": "Nouvelle Section", "items": [{"num": "1", "designation": "", "unite": "ff", "qte": 0, "pu": 0}]}); st.rerun()
                    st.session_state.devis_bat_main_oeuvre = st.number_input("👷 Main d'oeuvre", value=st.session_state.devis_bat_main_oeuvre, min_value=0.0, key="mo_devis_bat", format="%.2f")
                    cout_total_bat = total_general_bat + st.session_state.devis_bat_main_oeuvre; st.metric("COUT TOTAL DU PROJET", f"{cout_total_bat:,.2f} {devise_devis_bat}")
                    st.markdown("**Ingénieur: ESDRAS | Tél: +243 972 888 690 | Email: esdrastsangya@gmail.com**")
                    if st.button("📄 GÉNÉRER DEVIS PDF BÂTIMENT", type="primary", width="stretch", key="gen_devis_bat"):
                        if client_devis_bat and st.session_state.devis_bat_titre:
                            numero_devis = f"DEV-BAT-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                            try:
                                data_devis = {"numero": numero_devis, "type": "Bâtiment", "client": client_devis_bat, "telephone": tel_client_devis_bat, "titre": st.session_state.devis_bat_titre, "parcelle": parcelle_devis_bat, "localisation": localisation_devis_bat, "sections": st.session_state.devis_bat_sections, "main_oeuvre": st.session_state.devis_bat_main_oeuvre, "total": cout_total_bat, "devise": devise_devis_bat, "created_by": st.session_state.user_name, "created_at": datetime.now().isoformat()}
                                supabase.table('devis').insert(data_devis).execute(); st.success(f"✅ Devis enregistré : {numero_devis}")
                            except Exception as e: st.error("Erreur enregistrement"); st.code(repr(e)); st.stop()
                            pdf_bytes = generer_pdf_devis_consulting(numero_devis, "Bâtiment", client_devis_bat, st.session_state.devis_bat_titre, parcelle_devis_bat, localisation_devis_bat, st.session_state.devis_bat_sections, devise_devis_bat, tel_client_devis_bat, st.session_state.devis_bat_main_oeuvre, ing_nom_bat, ing_tel_bat)
                            st.session_state.pdf_devis_bat = pdf_bytes; st.session_state.num_devis_bat = numero_devis; st.rerun()
                        else: st.error("Client et Titre requis")
                    col_btn1, col_btn2, col_btn3 = st.columns(3)
                    with col_btn1:
                        if 'pdf_devis_bat' in st.session_state and st.session_state.pdf_devis_bat and peut_dl_bat:
                            st.download_button("📥 Télécharger PDF", data=st.session_state.pdf_devis_bat, file_name=f"{st.session_state.num_devis_bat}.pdf", mime="application/pdf", width="stretch", key="dl_devis_bat_1")
                    with col_btn2:
                        if 'pdf_devis_bat' in st.session_state and st.session_state.pdf_devis_bat and peut_pr_bat:
                            pdf_b64 = base64.b64encode(st.session_state.pdf_devis_bat).decode()
                            st.components.v1.html(f"""<button onclick="printPDF_bat()" style="width:100%; padding:10px; background:#00ff41; color:black; font-weight:bold; border:none; border-radius:5px; cursor:pointer;">🖨️ IMPRIMER LE DEVIS</button><script>function printPDF_bat(){{const pdfData='data:application/pdf;base64,{pdf_b64}';const win=window.open('','_blank');win.document.write('<iframe src="'+pdfData+'" width="100%" height="100%" style="border:none;"></iframe>');win.document.close();setTimeout(()=>{{win.print();}},1000);}}</script>""", height=60)
                    with col_btn3:
                        if st.button("🔄 Réinitialiser", key="reset_devis_bat_1", width="stretch"): st.session_state.devis_bat_sections = []
                        if 'pdf_devis_bat' in st.session_state: del st.session_state.pdf_devis_bat; st.rerun()
                else: st.info("🔒 Vous n'avez pas la permission de créer des devis bâtiment")
            tab_idx += 1

        # ===== 3. ONGLET FACTURE TRAVAUX BATIMENT =====
        if peut_facture_bat:
            with tabs[tab_idx]:
                st.subheader("🧾 Facture Travaux Exécutés - Bâtiment")
                st.info(f"**Ingénieur:** ESDRAS | **Tél:** +243 972 888 690 | **Email:** esdrastsangya@gmail.com")
                col1, col2, col3 = st.columns(3)
                with col1: client_fact_bat = st.text_input("👤 Client", key="client_fact_bat"); num_devis_ref = st.text_input("📄 N° Devis Référence", key="num_devis_ref_bat")
                with col2: st.session_state.titre_fact_bat = st.text_input("📋 Intitulé des Travaux", value=st.session_state.titre_fact_bat, key="titre_fact_bat_input"); date_fact_bat = st.date_input("📅 Date Facture", value=datetime.now().date(), key="date_fact_bat")
                with col3: st.session_state.facture_bat_pourcentage = st.number_input("📊 % Travaux Exécutés", min_value=0.0, max_value=100.0, value=st.session_state.facture_bat_pourcentage, key="pourc_fact_bat", format="%.2f"); devise_fact_bat = st.selectbox("💵 Devise", ["USD", "FC", "€"], key="devise_fact_bat")
                st.divider()
                if not st.session_state.facture_bat_sections: st.session_state.facture_bat_sections = [{"numero": "I", "titre": "Installation chantier / Demolitions", "items": [{"num": "1", "designation": "Installation chantier", "unite": "ff", "qte_totale": 1, "pu": 200, "qte_execute": 1}]},{"numero": "II", "titre": "fondation", "items": [{"num": "1", "designation": "moellon", "unite": "Canters", "qte_totale": 9, "pu": 50, "qte_execute": 0}]}]
                total_facture = 0
                for idx, section in enumerate(st.session_state.facture_bat_sections):
                    col_titre1, col_titre2 = st.columns([0.2, 3])
                    with col_titre1: section['numero'] = st.text_input("N°Sec", value=section['numero'], key=f"numsec_fact_{idx}", label_visibility="collapsed")
                    with col_titre2: section['titre'] = st.text_input("Titre Section", value=section['titre'], key=f"titresec_fact_{idx}", label_visibility="collapsed")
                    st.markdown(f"**{section['numero']}. {section['titre']}**")
                    sous_total_sec = 0
                    items_to_delete = []
                    for i, item in enumerate(section['items']):
                        col1, col2, col3, col4, col5, col6, col7, col8 = st.columns([0.5, 3, 1, 1.2, 1.2, 1.2, 1.2, 0.5])
                        with col1: st.write(item.get('num',''))
                        with col2: st.write(item.get('designation',''))
                        with col3: st.write(item.get('unite',''))
                        with col4: st.write(f"{item.get('qte_totale',0)}")
                        with col5: st.write(f"{item.get('pu',0):,.2f}")
                        with col6: qte_ex = st.number_input("Qté Exécutée", value=float(item.get('qte_execute',0)), min_value=0.0, key=f"qte_ex_bat_{idx}_{i}", label_visibility="collapsed", format="%.2f"); section['items'][i]['qte_execute'] = qte_ex
                        with col7: montant = qte_ex * item.get('pu',0); st.markdown(f"**{montant:,.2f}**"); sous_total_sec += montant
                        with col8:
                            if st.button("❌", key=f"del_fact_bat_{idx}_{i}"): items_to_delete.append(i)
                    for i in sorted(items_to_delete, reverse=True): section['items'].pop(i); st.rerun()
                    total_facture += sous_total_sec; st.markdown(f"**Sous-total Exécuté: {sous_total_sec:,.2f}**")
                    if st.button("➕ Ajouter Ligne", key=f"add_line_fact_{idx}"): section['items'].append({"num": "", "designation": "", "unite": "ff", "qte_totale": 0, "pu": 0, "qte_execute": 0}); st.rerun()
                    st.divider()
                if st.button("➕ Ajouter Section Facture", key="add_section_fact", width="stretch"):
                    new_num = f"Sec{len(st.session_state.facture_bat_sections)+1}"
                    st.session_state.facture_bat_sections.append({"numero": new_num, "titre": "Nouvelle Section", "items": [{"num": "1", "designation": "", "unite": "ff", "qte_totale": 0, "pu": 0, "qte_execute": 0}]}); st.rerun()
                col_mo1, col_mo2, col_mo3 = st.columns(3)
                with col_mo1: retenue = st.number_input("💰 Retenue Garantie %", value=5.0, min_value=0.0, max_value=100.0, key="retenue_fact_bat")
                with col_mo2: montant_retenue = total_facture * (retenue/100); st.metric("MONTANT RETENUE", f"{montant_retenue:,.2f} {devise_fact_bat}")
                with col_mo3: net_a_payer = total_facture - montant_retenue; st.metric("NET A PAYER", f"{net_a_payer:,.2f} {devise_fact_bat}")
                st.markdown("**Ingénieur: ESDRAS | Tél: +243 972 888 690 | Email: esdrastsangya@gmail.com**")
                if st.button("📄 GÉNÉRER FACTURE PDF", type="primary", width="stretch", key="gen_fact_bat"):
                    if client_fact_bat and st.session_state.titre_fact_bat:
                        numero_fact = f"FACT-BAT-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                        try:
                            data_fact = {"numero": numero_fact, "type": "Facture Bâtiment", "client": client_fact_bat, "titre": st.session_state.titre_fact_bat, "date": str(date_fact_bat), "num_devis_ref": num_devis_ref, "sections": st.session_state.facture_bat_sections, "pourcentage": st.session_state.facture_bat_pourcentage, "retenue": retenue, "total": total_facture, "net_a_payer": net_a_payer, "devise": devise_fact_bat, "created_by": st.session_state.user_name, "created_at": datetime.now().isoformat()}
                            supabase.table('factures').insert(data_fact).execute(); st.success(f"✅ Facture enregistrée : {numero_fact}")
                        except Exception as e: st.error("Erreur enregistrement"); st.code(repr(e)); st.stop()
                        pdf_bytes = generer_pdf_facture_consulting(numero_fact, client_fact_bat, st.session_state.titre_fact_bat, date_fact_bat, num_devis_ref, st.session_state.facture_bat_sections, devise_fact_bat, total_facture, montant_retenue, net_a_payer, "ESDRAS", "+243 972 888 690")
                        st.session_state.pdf_fact_bat = pdf_bytes; st.session_state.num_fact_bat = numero_fact; st.rerun()
                    else: st.error("Client et Intitulé des travaux requis")
                if 'pdf_fact_bat' in st.session_state and st.session_state.pdf_fact_bat: st.download_button("📥 Télécharger Facture PDF", data=st.session_state.pdf_fact_bat, file_name=f"{st.session_state.num_fact_bat}.pdf", mime="application/pdf", width="stretch", key="dl_fact_bat_1")
            tab_idx += 1

        # ===== 4. ONGLET HISTORIQUE =====
        if (peut_hist_ind or peut_hist_bat):
            with tabs[tab_idx]:
                st.subheader("📜 Historique")
                hist_tabs_list = []
                if peut_hist_ind: hist_tabs_list.append("🏭 Devis Industriels")
                if peut_hist_bat: hist_tabs_list.append("🏗️ Devis + Factures Bâtiment")
                hist_tabs = st.tabs(hist_tabs_list); hist_idx = 0
                if peut_hist_ind:
                    with hist_tabs[hist_idx]:
                        try: devis_list = supabase.table('devis').select("*").eq('type', 'Industriel').order("created_at", desc=True).limit(20).execute().data
                        except: devis_list = []
                        if not devis_list: st.info("Aucun devis industriel enregistré")
                        else:
                            for d in devis_list:
                                with st.expander(f"🏭 {d.get('numero')} - {d.get('client')} - {d.get('total',0):,.0f} {d.get('devise','USD')}"):
                                    st.write(f"**Titre:** {d.get('titre')}"); st.write(f"**Créé par:** {d.get('created_by')} le {str(d.get('created_at'))[:10]}"); st.write(f"**Localisation:** {d.get('localisation','')}")
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        if peut_dl_ind:
                                            sections_data = d.get('sections');
                                            if isinstance(sections_data, str): sections_data = json.loads(sections_data)
                                            pdf_bytes = generer_pdf_devis_consulting(d.get('numero'), "Industriel", d.get('client'), d.get('titre'), d.get('parcelle'), d.get('localisation'), sections_data, d.get('devise'), d.get('telephone'), d.get('main_oeuvre'), "SAMY TSANGYA", "+256766515428")
                                            st.download_button("📥 Télécharger PDF", data=pdf_bytes, file_name=f"{d.get('numero')}.pdf", mime="application/pdf", key=f"dl_hist_ind_{d.get('numero')}", width="stretch")
                                    with col2:
                                        if peut_pr_ind:
                                            sections_data = d.get('sections');
                                            if isinstance(sections_data, str): sections_data = json.loads(sections_data)
                                            pdf_bytes = generer_pdf_devis_consulting(d.get('numero'), "Industriel", d.get('client'), d.get('titre'), d.get('parcelle'), d.get('localisation'), sections_data, d.get('devise'), d.get('telephone'), d.get('main_oeuvre'), "SAMY TSANGYA", "+256766515428")
                                            pdf_b64 = base64.b64encode(pdf_bytes).decode(); safe_id = str(d.get('numero','DEV')).replace('-', '_')
                                            st.components.v1.html(f"""<button onclick="printPDF_ind_{safe_id}()" style="width:100%; padding:10px; background:#00ff41; color:black; font-weight:bold; border:none; border-radius:5px; cursor:pointer;">🖨️ IMPRIMER</button><script>function printPDF_ind_{safe_id}() {{const pdfData = 'data:application/pdf;base64,{pdf_b64}';const win = window.open('', '_blank');win.document.write('<iframe src="'+pdfData+'" width="100%" height="100%" style="border:none;"></iframe>');win.document.close();setTimeout(()=>{{win.print();}},1000);}}</script>""", height=60)
                    hist_idx += 1
                if peut_hist_bat:
                    with hist_tabs[hist_idx]:
                        sub_hist_tabs = st.tabs(["📄 Devis Bâtiment", "🧾 Factures Bâtiment"])
                        with sub_hist_tabs[0]:
                            try: devis_list_bat = supabase.table('devis').select("*").eq('type', 'Bâtiment').order("created_at", desc=True).limit(20).execute().data
                            except: devis_list_bat = []
                            if not devis_list_bat: st.info("Aucun devis bâtiment enregistré")
                            else:
                                for d in devis_list_bat:
    with st.expander(f"🏗️ {d.get('numero')} - {d.get('client')} - {d.get('total',0):,.0f} {d.get('devise','USD')}"):
        st.write(f"**Titre:** {d.get('titre')}")
        st.write(f"**Créé par:** {d.get('created_by')} le {str(d.get('created_at'))[:10]}")
        st.write(f"**Localisation:** {d.get('localisation','')}")
        st.write(f"**Ingénieur:** ESDRAS | **Tél:** +243 972 888 690")

        col1, col2 = st.columns(2)
        with col1:
            if peut_dl_bat:
                sections_data = d.get('sections')
                if isinstance(sections_data, str):
                    sections_data = json.loads(sections_data)
                pdf_bytes = generer_pdf_devis_consulting(
                    d.get('numero'), "Bâtiment", d.get('client'), d.get('titre'),
                    d.get('parcelle'), d.get('localisation'), sections_data,
                    d.get('devise'), d.get('telephone'), d.get('main_oeuvre'), "ESDRAS", "+243 972 888 690"
                )
                st.download_button(
                    "📥 Télécharger PDF",
                    data=pdf_bytes,
                    file_name=f"{d.get('numero')}.pdf",
                    mime="application/pdf",
                    key=f"dl_hist_bat_devis_{d.get('numero')}",
                    width="stretch"
                )
        with col2:
            if peut_pr_bat:
                sections_data = d.get('sections')
                if isinstance(sections_data, str):
                    sections_data = json.loads(sections_data)

                pdf_bytes = generer_pdf_devis_consulting(
                    d.get('numero'), "Bâtiment", d.get('client'), d.get('titre'),
                    d.get('parcelle'), d.get('localisation'), sections_data,
                    d.get('devise'), d.get('telephone'), d.get('main_oeuvre'), "ESDRAS", "+243 972 888 690"
                )
                pdf_b64 = base64.b64encode(pdf_bytes).decode()
                safe_id = str(d.get('numero','DEV')).replace('-', '_')

                st.components.v1.html(f"""<button onclick="printPDF_bat_devis_{safe_id}()" style="width:100%; padding:10px; background:#00ff41; color:black; font-weight:bold; border:none; border-radius:5px; cursor:pointer;">
                    🖨️ IMPRIMER
                </button>
                <script>
                function printPDF_bat_devis_{safe_id}() {{
                    const pdfData = 'data:application/pdf;base64,{pdf_b64}';
                    const win = window.open('', '_blank');
                    win.document.write('<iframe src="'+pdfData+'" width="100%" height="100%" style="border:none;"></iframe>');
                    win.document.close();
                    setTimeout(()=>{{win.print();}},1000);
                }}
                </script>
            """, height=60)

# ===== 2. FACTURES BATIMENT =====
with sub_hist_tabs[1]:
    try:
        factures_list_bat = supabase.table('factures').select("*").eq('type', 'Facture Bâtiment').order("created_at", desc=True).limit(20).execute().data
    except:
        factures_list_bat = []

    if not factures_list_bat:
        st.info("Aucune facture bâtiment enregistrée")
    else:
        for f in factures_list_bat:
            with st.expander(f"🧾 {f.get('numero')} - {f.get('client')} - NET: {f.get('net_a_payer',0):,.0f} {f.get('devise','USD')}"):
                st.write(f"**Intitulé:** {f.get('titre')}")
                st.write(f"**Devis Ref:** {f.get('num_devis_ref','')}")
                st.write(f"**Créé par:** {f.get('created_by')} le {str(f.get('created_at'))[:10]}")
                st.write(f"**% Exécuté:** {f.get('pourcentage',0)}% | **Retenue:** {f.get('retenue',0)}%")
                st.write(f"**Ingénieur:** ESDRAS | **Tél:** +243 972 888 690")

                col1, col2 = st.columns(2)
                with col1:
                    if peut_dl_bat:
                        sections_data = f.get('sections')
                        if isinstance(sections_data, str):
                            sections_data = json.loads(sections_data)
                        pdf_bytes = generer_pdf_facture_consulting(
                            f.get('numero'), f.get('client'), f.get('titre'), f.get('date'), f.get('num_devis_ref'),
                            sections_data, f.get('devise'), f.get('total'),
                            f.get('total',0) * (f.get('retenue',0)/100), f.get('net_a_payer'), "ESDRAS", "+243 972 888 690"
                        )
                        st.download_button(
                            "📥 Télécharger Facture",
                            data=pdf_bytes,
                            file_name=f"{f.get('numero')}.pdf",
                            mime="application/pdf",
                            key=f"dl_hist_bat_fact_{f.get('numero')}",
                            width="stretch"
                        )
                with col2:
                    if peut_pr_bat:
                        sections_data = f.get('sections')
                        if isinstance(sections_data, str):
                            sections_data = json.loads(sections_data)
                        pdf_bytes = generer_pdf_facture_consulting(
                            f.get('numero'), f.get('client'), f.get('titre'), f.get('date'), f.get('num_devis_ref'),
                            sections_data, f.get('devise'), f.get('total'),
                            f.get('total',0) * (f.get('retenue',0)/100), f.get('net_a_payer'), "ESDRAS", "+243 972 888 690"
                        )
                        pdf_b64 = base64.b64encode(pdf_bytes).decode()
                        safe_id = str(f.get('numero','FACT')).replace('-', '_')

                        st.components.v1.html(f"""<button onclick="printPDF_bat_fact_{safe_id}()" style="width:100%; padding:10px; background:#ff9500; color:white; font-weight:bold; border:none; border-radius:5px; cursor:pointer;">
                            🖨️ IMPRIMER FACTURE
                        </button>
                        <script>
                        function printPDF_bat_fact_{safe_id}() {{
                            const pdfData = 'data:application/pdf;base64,{pdf_b64}';
                            const win = window.open('', '_blank');
                            win.document.write('<iframe src="'+pdfData+'" width="100%" height="100%" style="border:none;"></iframe>');
                            win.document.close();
                            setTimeout(()=>{{win.print();}},1000);
                        }}
                        </script>
                    """, height=60)

hist_idx += 1

                        
if "👥 Utilisateurs" in tab_map:
    with tab_map["👥 Utilisateurs"]:
        st.markdown("## 👥 Gestion Utilisateurs - Droits d'Accès")
        
        df_utilisateurs = load_table("utilisateurs")
        df_compta = load_table("compta")
        
        # === 1. AJOUTER NOUVEL UTILISATEUR ===
        with st.expander("➕ Ajouter Nouvel Utilisateur", expanded=False):
            with st.form("form_user", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                nom_user = c1.text_input("Nom *", placeholder="Ex: Jean KABAMBA")
                role_user = c2.selectbox("Rôle *", ["PDG", "GERANTE", "UTILISATEUR", "CAISSIER", "COMMERCIAL"])
                pwd_user = c3.text_input("Mot de passe *", type="password")
                
                st.markdown("**🔐 Autorisations d'onglets :**")
                col1, col2, col3, col4 = st.columns(4)
                perm_dashboard = col1.checkbox("Dashboard", value=True)
                perm_commerce = col2.checkbox("Commerce", value=True)
                perm_stock = col3.checkbox("Gestion Stock")
                perm_immobilier = col4.checkbox("Immobilier")
                perm_automobile = col1.checkbox("Automobile")
                perm_parc = col2.checkbox("Gestion Parc")
                perm_comptabilite = col3.checkbox("Comptabilité")
                perm_factures = col4.checkbox("Factures")
                perm_supprimer = col1.checkbox("🗑️ Peut Supprimer")
                perm_users = col2.checkbox("👥 Gérer Utilisateurs")

                st.markdown("**📋 Autorisations Devis :**")
                col_d1, col_d2, col_d3 = st.columns(3)
                with col_d1:
                    st.markdown("*Devis Industriel*")
                    perm_devis_ind = st.checkbox("Créer", key="perm_ind_creer")
                    perm_devis_ind_dl = st.checkbox("Télécharger", key="perm_ind_dl")
                    perm_devis_ind_pr = st.checkbox("Imprimer", key="perm_ind_pr")
                with col_d2:
                    st.markdown("*Devis Bâtiment*")
                    perm_devis_bat = st.checkbox("Créer", key="perm_bat_creer")
                    perm_devis_bat_dl = st.checkbox("Télécharger", key="perm_bat_dl")
                    perm_devis_bat_pr = st.checkbox("Imprimer", key="perm_bat_pr")
                with col_d3:
                    st.markdown("*Historique*")
                    perm_devis_hist = st.checkbox("Voir Historique", key="perm_hist")

                st.markdown("**📂 Catégories de Factures Visibles :**")
                cats_dispo = sorted(df_compta['categorie'].dropna().unique().tolist()) if 'categorie' in df_compta.columns else []
                cats_autorisees = st.multiselect("Sélectionne les catégories", ["Toutes"] + cats_dispo, default=["Toutes"], key="cats_factures")

                if st.form_submit_button("💾 Ajouter Utilisateur", type="primary"):
                    if nom_user and pwd_user:
                        try:
                            perms_dict = {
                                "dashboard": perm_dashboard, "commerce": perm_commerce, "stock": perm_stock,
                                "immobilier": perm_immobilier, "automobile": perm_automobile, "parc": perm_parc,
                                "comptabilite": perm_comptabilite, "factures": perm_factures, "supprimer": perm_supprimer,
                                "users": perm_users, "devis_industriel": perm_devis_ind,
                                "devis_industriel_download": perm_devis_ind_dl, "devis_industriel_print": perm_devis_ind_pr,
                                "devis_batiment": perm_devis_bat, "devis_batiment_download": perm_devis_bat_dl,
                                "devis_batiment_print": perm_devis_bat_pr, "devis_historique": perm_devis_hist
                            }
                            supabase.table("utilisateurs").insert({
                                "nom": nom_user.upper(),
                                "role": role_user,
                                "password": pwd_user,
                                "permissions": perms_dict,
                                "categories_autorisees": [] if "Toutes" in cats_autorisees else cats_autorisees
                            }).execute()
                            st.success(f"✅ Utilisateur {nom_user} ajouté")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error("Erreur ajout")
                            st.code(repr(e))
                    else:
                        st.error("Nom et mot de passe obligatoires")

        st.divider()
        
        # === 2. LISTE ET MODIFICATION DES UTILISATEURS ===
        st.subheader("📋 Liste des Utilisateurs")
        if df_utilisateurs.empty:
            st.info("Aucun utilisateur")
        else:
            for _, user in df_utilisateurs.iterrows():
                current_perms = user.get('permissions', {})
                if isinstance(current_perms, str):
                    try: current_perms = json.loads(current_perms)
                    except: current_perms = {}

                with st.expander(f"{user['nom']} - {user['role']}"):
                    # AFFICHAGE ACTUEL
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        st.write("**Onglets :**")
                        for k,v in [("dashboard","Dashboard"),("commerce","Commerce"),("stock","Stock"),("immobilier","Immobilier"),("automobile","Automobile"),("parc","Parc"),("comptabilite","Comptabilité"),("factures","Factures"),("users","Utilisateurs"),("supprimer","Supprimer")]:
                            if current_perms.get(k): st.write(f"✅ {v}")
                    with c2:
                        st.write("**Devis Industriel :**")
                        if current_perms.get('devis_industriel'): st.write("✅ Créer")
                        if current_perms.get('devis_industriel_download'): st.write("✅ Télécharger")
                        if current_perms.get('devis_industriel_print'): st.write("✅ Imprimer")
                        if current_perms.get('devis_historique_industriel'): st.write("✅ Historique")
                    with c3:
                        st.write("**Devis Bâtiment :**")
                        if current_perms.get('devis_batiment'): st.write("✅ Créer")
                        if current_perms.get('devis_batiment_download'): st.write("✅ Télécharger")
                        if current_perms.get('devis_batiment_print'): st.write("✅ Imprimer")
                        if current_perms.get('devis_historique_batiment'): st.write("✅ Historique")

                    st.divider()
                    
                    # SEUL LE PDG PEUT MODIFIER
                    if st.session_state.user_role == "PDG":
                        
                        tab1, tab2 = st.tabs(["✏️ Permissions", "🔑 Mot de passe"])
                        
                        # TAB 1: PERMISSIONS
                        with tab1:
                            with st.form(f"edit_user_{user['id']}"):
                                col1, col2, col3, col4 = st.columns(4)
                                perm_dashboard = col1.checkbox("Dashboard", value=current_perms.get('dashboard', False), key=f"edit_dash_{user['id']}")
                                perm_commerce = col2.checkbox("Commerce", value=current_perms.get('commerce', False), key=f"edit_com_{user['id']}")
                                perm_stock = col3.checkbox("Gestion Stock", value=current_perms.get('stock', False), key=f"edit_stock_{user['id']}")
                                perm_immobilier = col4.checkbox("Immobilier", value=current_perms.get('immobilier', False), key=f"edit_immo_{user['id']}")
                                perm_automobile = col1.checkbox("Automobile", value=current_perms.get('automobile', False), key=f"edit_auto_{user['id']}")
                                perm_parc = col2.checkbox("Gestion Parc", value=current_perms.get('parc', False), key=f"edit_parc_{user['id']}")
                                perm_comptabilite = col3.checkbox("Comptabilité", value=current_perms.get('comptabilite', False), key=f"edit_comp_{user['id']}")
                                perm_factures = col4.checkbox("Factures", value=current_perms.get('factures', False), key=f"edit_fact_{user['id']}")
                                perm_supprimer = col1.checkbox("🗑️ Peut Supprimer", value=current_perms.get('supprimer', False), key=f"edit_sup_{user['id']}")
                                perm_users = col2.checkbox("👥 Gérer Utilisateurs", value=current_perms.get('users', False), key=f"edit_users_{user['id']}")

                                st.markdown("**📋 Devis Industriel :**")
                                col_i1, col_i2, col_i3, col_i4 = st.columns(4)
                                perm_devis_ind = col_i1.checkbox("Créer", value=current_perms.get('devis_industriel', False), key=f"edit_ind_{user['id']}")
                                perm_devis_ind_dl = col_i2.checkbox("Télécharger", value=current_perms.get('devis_industriel_download', False), key=f"edit_ind_dl_{user['id']}")
                                perm_devis_ind_pr = col_i3.checkbox("Imprimer", value=current_perms.get('devis_industriel_print', False), key=f"edit_ind_pr_{user['id']}")
                                perm_devis_hist_ind = col_i4.checkbox("Historique", value=current_perms.get('devis_historique_industriel', False), key=f"edit_hist_ind_{user['id']}") # <-- NOUVEAU

                                st.markdown("**📋 Devis Bâtiment :**")
                                col_b1, col_b2, col_b3, col_b4 = st.columns(4)
                                perm_devis_bat = col_b1.checkbox("Créer", value=current_perms.get('devis_batiment', False), key=f"edit_bat_{user['id']}")
                                perm_devis_bat_dl = col_b2.checkbox("Télécharger", value=current_perms.get('devis_batiment_download', False), key=f"edit_bat_dl_{user['id']}")
                                perm_devis_bat_pr = col_b3.checkbox("Imprimer", value=current_perms.get('devis_batiment_print', False), key=f"edit_bat_pr_{user['id']}")
                                perm_devis_hist_bat = col_b4.checkbox("Historique", value=current_perms.get('devis_historique_batiment', False), key=f"edit_hist_bat_{user['id']}") # <-- NOUVEAU

                                if st.form_submit_button("💾 Enregistrer Permissions", type="primary", width="stretch"):
                                    new_perms = {
                                        "dashboard": perm_dashboard, "commerce": perm_commerce, "stock": perm_stock,
                                        "immobilier": perm_immobilier, "automobile": perm_automobile, "parc": perm_parc,
                                        "comptabilite": perm_comptabilite, "factures": perm_factures, "supprimer": perm_supprimer,
                                        "users": perm_users, 
                                        "devis_industriel": perm_devis_ind,
                                        "devis_industriel_download": perm_devis_ind_dl, 
                                        "devis_industriel_print": perm_devis_ind_pr,
                                        "devis_historique_industriel": perm_devis_hist_ind, # <-- NOUVEAU
                                        "devis_batiment": perm_devis_bat, 
                                        "devis_batiment_download": perm_devis_bat_dl,
                                        "devis_batiment_print": perm_devis_bat_pr, 
                                        "devis_historique_batiment": perm_devis_hist_bat # <-- NOUVEAU
                                    }
                                    try:
                                        supabase.table("utilisateurs").update({"permissions": new_perms}).eq("id", int(user['id'])).execute()
                                        st.success(f"Permissions de {user['nom']} mises à jour")
                                        st.cache_data.clear()
                                        st.rerun()
                                    except Exception as e:
                                        st.error("Erreur modification")
                                        st.code(repr(e))

                        # TAB 2: MOT DE PASSE
                        with tab2:
                            with st.form(f"pwd_user_{user['id']}"):
                                new_pwd = st.text_input("Nouveau mot de passe", type="password", key=f"newpwd_{user['id']}")
                                confirm_pwd = st.text_input("Confirmer mot de passe", type="password", key=f"confpwd_{user['id']}")
                                if st.form_submit_button("🔐 Mettre à jour mot de passe"):
                                    if new_pwd and new_pwd == confirm_pwd:
                                        if len(new_pwd) < 4:
                                            st.warning("Min 4 caractères")
                                        else:
                                            try:
                                                supabase.table("utilisateurs").update({"password": new_pwd}).eq("id", int(user['id'])).execute()
                                                st.success(f"✅ Mot de passe de {user['nom']} changé")
                                                st.rerun()
                                            except Exception as e:
                                                st.error("Erreur MDP")
                                                st.code(repr(e))
                                    else:
                                        st.error("Les mots de passe ne correspondent pas")

                        st.divider()
                        # SUPPRIMER
                        if user['nom'] != st.session_state.user_name:
                            if st.button("🗑️ Supprimer cet utilisateur", key=f"del_user_{user['id']}", type="secondary", width="stretch"):
                                try:
                                    supabase.table("utilisateurs").delete().eq("id", int(user['id'])).execute()
                                    st.success(f"Utilisateur {user['nom']} supprimé")
                                    st.cache_data.clear()
                                    st.rerun()
                                except Exception as e:
                                    st.error("Erreur suppression")
                                    st.code(repr(e))
                        else:
                            st.info("🔒 Vous ne pouvez pas supprimer votre propre compte")
                    else:
                        st.info("🔒 Seul le PDG peut modifier")
                         # === FLOKI SOLDAT COMPLET - VERSION PDG ===
import difflib
import re
import urllib.parse
import json
import requests
import streamlit as st
import pandas as pd
from datetime import datetime

class FLOKI:
    def __init__(self, supabase_client, dataframes):
        self.supabase = supabase_client
        self.df = dataframes
        self.system_knowledge = self._get_supabase_schema()

    def _get_supabase_schema(self):
        schema = {}
        tables = ["articles", "compta", "biens", "voitures", "mouvements_stock", "devis", "notifications", "floki_logs"]
        for t in tables:
            try:
                result = self.supabase.table(t).select("*").limit(1).execute()
                schema[t] = list(result.data[0].keys()) if result.data else []
            except:
                schema[t] = []
        return schema

    def ask(self, question):
        q = question.lower().strip()
        log_entry = {"demande": question, "date": datetime.now().isoformat(), "source": "ASYMAS"}

        if any(g in q for g in ["slt", "salut", "bonjour", "hello", "yo"]):
            return "Présent chef. FLOKI opérationnel. Donnez l'ordre."

        if "envoie" in q and "message" in q and "numero" in q:
            result = self._action_send_whatsapp(question)
            log_entry.update({"action": "whatsapp_send", "reponse": result})
            self._log_action(log_entry)
            return result

        if any(k in q for k in ["redige", "rédige", "lettre", "relance", "convocation"]):
            result = self._action_rediger(question)
            log_entry.update({"action": "redaction", "reponse": result})
            self._log_action(log_entry)
            return result

        if any(k in q for k in ["conseil", "avis", "opportunite", "risque", "que faire"]):
            result = self._action_conseil(q)
            log_entry.update({"action": "conseil", "reponse": result})
            self._log_action(log_entry)
            return result

        q_clean = re.sub(r'(trouve moi|donne moi|donne|trouve|cherche|le prix de|prix du|du|de|le|la|un|une|pour moi|combien)', '', q).strip()
        rep = self._search_asymas(q_clean)
        if rep:
            log_entry.update({"source": "ASYMAS", "reponse": rep})
            self._log_action(log_entry)
            return rep + "\n\nSource: ASYMAS"

        web_rep = self._search_web(question)
        log_entry.update({"source": "WEB", "reponse": web_rep})
        self._log_action(log_entry)
        return web_rep + "\n\nSource: WEB"

    def _search_asymas(self, q):
        # Voiture moins chère
        if "voiture" in q and ("moins cher" in q or "prix" in q):
            return self._get_voiture_moins_cher()

        # Liste voitures
        if "voiture" in q and ("liste" in q or "donne" in q):
            return self._get_voitures_stock()

        # Produit
        rep = self._search_product(q)
        if rep: return rep

        # Pertes commerce
        if "perte" in q and "commerce" in q:
            return self._get_pertes_commerce()

        # Stock bas
        if any(k in q for k in ["stock bas", "rupture", "manque"]):
            return self._stock_bas()

        # CA
        if any(k in q for k in ["ca", "chiffre", "revenu", "vente", "argent", "benefice", "solde"]):
            return self._chiffre_affaires()

        return None

    def _get_voiture_moins_cher(self):
        if self.df['voitures'].empty:
            return "Pas de données voitures chef."
        prix_col = next((col for col in ['prix', 'prix_vente', 'prix_achat', 'montant'] if col in self.df['voitures'].columns), None)
        if not prix_col:
            return "Chef, je ne trouve pas la colonne prix dans voitures."
        dispo = self.df['voitures'][self.df['voitures'].get('quantite', 1) > 0]
        if dispo.empty:
            return "Aucune voiture en stock chef."
        moins_chere = dispo.loc[dispo[prix_col].idxmin()]
        modele = moins_chere.get('modele', moins_chere.get('nom', 'N/A'))
        prix = float(moins_chere[prix_col])
        return f"Voiture la moins chère: {modele} à {prix:,.0f} FC"

    def _get_voitures_stock(self):
        if self.df['voitures'].empty:
            return "Pas de données voitures chef."
        dispo = self.df['voitures'][self.df['voitures'].get('quantite', 0) > 0]
        if dispo.empty:
            return "Aucune voiture en stock chef."
        txt = "\n".join([f"- {r.get('modele', r.get('nom', 'N/A'))}: {int(r.get('quantite', 0))} unités - {float(r.get('prix', r.get('prix_vente', 0))):,.0f} FC" for _, r in dispo.iterrows()])
        return f"Voitures en stock:\n{txt}"

    def _get_pertes_commerce(self):
        try:
            result = self.supabase.table("mouvements_stock").select("*").eq("type", "perte").eq("categorie", "commerce").order("date", desc=True).limit(10).execute()
            if not result.data:
                return "Aucune perte commerce enregistrée chef."
            txt = "\n".join([f"- {r.get('article', 'N/A')}: {r.get('montant', 0):,.0f} FC le {r.get('date', '')[:10]}" for r in result.data])
            return f"Dernières pertes commerce:\n{txt}"
        except Exception as e:
            return f"Erreur lecture pertes: {e}. Vérifiez RLS sur mouvements_stock."

    def _search_product(self, q):
        if self.df['articles'].empty:
            return None
        articles = self.df['articles'].copy()
        articles['nom_clean'] = articles['nom_article'].astype(str).str.lower().str.replace(r'[^a-z0-9\s]', '', regex=True).str.replace(r'\s+', ' ', regex=True).str.strip()
        q_clean = re.sub(r'[^a-z0-9\s]', '', q).strip()
        mots_q = [w for w in q_clean.split() if len(w) > 2]
        if mots_q:
            for _, r in articles.iterrows():
                if all(word in r['nom_clean'] for word in mots_q):
                    return f"{r['nom_article']}: Stock {int(r['stock'])} unités, Prix {float(r['prix_vente']):,.0f} FC"
        noms = articles['nom_clean'].tolist()
        closest = difflib.get_close_matches(q_clean, noms, n=1, cutoff=0.45)
        if closest:
            r = articles[articles['nom_clean'] == closest[0]].iloc[0]
            return f"{r['nom_article']}: Stock {int(r['stock'])} unités, Prix {float(r['prix_vente']):,.0f} FC"
        return None

    def _stock_bas(self):
        if self.df['articles'].empty:
            return "Pas d'articles chef."
        low = self.df['articles'][self.df['articles']['stock'] < 5]
        if low.empty:
            return "Stock OK chef. Rien en dessous de 5 unités."
        txt = "\n".join([f"- {r['nom_article']}: {r['stock']} unités" for _, r in low.iterrows()])
        return f"Attention chef, stock bas:\n{txt}"

    def _chiffre_affaires(self):
        if self.df['compta'].empty:
            return "Pas de données compta chef."
        rev = self.df['compta'][self.df['compta']['type'] == 'Revenu']['montant'].sum()
        dep = self.df['compta'][self.df['compta']['type'] == 'Dépense']['montant'].sum()
        return f"Rapport compta:\nRevenus: {rev:,.0f} FC\nDépenses: {dep:,.0f} FC\nSolde: {rev-dep:,.0f} FC"

    def _search_web(self, q):
        try:
            url = f"https://api.duckgo.com/?q={urllib.parse.quote(q)}&format=json&no_html=1"
            r = requests.get(url, timeout=4)
            data = r.json()
            if data.get('AbstractText'):
                return f"Info vérifiée: {data['AbstractText']}"
            return f"Négatif chef. Rien de vérifiable sur le web pour '{q}'."
        except:
            return "Le web ne répond pas chef."

    def _action_rediger(self, question):
        if "relance" in question.lower():
            return "Objet: Relance de paiement\nMonsieur/Madame,\n\nNous constatons que la facture reste impayée.\nMerci de régulariser sous 48h.\n\nASYMAS BUSINESS"
        if "convocation" in question.lower():
            return "Objet: Convocation\nVous êtes convoqué(e) le [DATE] à [HEURE] pour [OBJET].\n\nASYMAS BUSINESS"
        return "Chef, précise: 'redige une relance' ou 'redige une convocation'."

    def _action_send_whatsapp(self, question):
        nums = re.findall(r'\+?\d{9,15}', question)
        if not nums:
            return "Chef, donne-moi un numéro. Ex: 'envoie un message au +243995105623 salut'"
        numero = nums[0].replace("+", "")
        message = re.sub(r'envoie un message.*?\+?\d{9,15}', '', question).strip() or "Message de ASYMAS BUSINESS"
        url = f"https://wa.me/{numero}?text={urllib.parse.quote(message)}"
        return f"Lien WhatsApp prêt: {url}"

    def notify_internal(self, message):
        try:
            self.supabase.table("notifications").insert({
                "message": f"[{st.session_state.get('user_name', 'PDG')}]: {message}",
                "created_at": datetime.now().isoformat()
            }).execute()
            return "Notification envoyée à l’équipe chef."
        except Exception as e:
            return f"Échec notification: {e}"

    def _action_conseil(self, q):
        if not self.df['articles'].empty and not self.df['compta'].empty:
            stock_bas = len(self.df['articles'][self.df['articles']['stock'] < 5])
            rev = self.df['compta'][self.df['compta']['type'] == 'Revenu']['montant'].sum()
            return f"FAIT: {stock_bas} articles en stock bas. CA: {rev:,.0f} FC.\nCONSEIL: Réapprovisionne sous 48h.\nRISQUE: Rupture = perte de vente."
        return "Chef, je croise vos données ASYMAS pour donner fait, conseil, risque."

    def _log_action(self, log_entry):
        try:
            self.supabase.table("floki_logs").insert(log_entry).execute()
        except:
            pass

# === UI FLOKI ===
if 'floki' not in st.session_state:
    dataframes = {
        "articles": df_articles,
        "compta": df_compta,
        "biens": df_biens,
        "voitures": df_voitures
    }
    st.session_state.floki = FLOKI(supabase, dataframes)

with st.sidebar:
    st.divider()
    st.markdown("### 🤖 FLOKI")
    st.caption("Conseiller du PDG - Comprend le système ASYMAS")

    q = st.text_input("Ordre pour FLOKI", key="floki_input",
                      placeholder="Ex: liste de mes voitures, voiture moins cher, CA du mois")

    st.info("🎤 Micro désactivé temporairement. Utilisez Chrome + localhost pour l'activer plus tard.")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Exécuter", type="primary", use_container_width=True):
            if q:
                with st.spinner("FLOKI réfléchit..."):
                    rep = st.session_state.floki.ask(q)
                    st.session_state.floki_rep = rep

    with col2:
        if st.button("Notifier équipe", use_container_width=True):
            if q:
                msg = st.session_state.floki.notify_internal(q)
                st.toast(msg)

    if 'floki_rep' in st.session_state:
        rep_clean = st.session_state.floki_rep.replace('"', '\\"').replace("\n", " ").replace("'", "\\'")
        st.components.v1.html(f"""
            <script>
            if ('speechSynthesis' in window) {{
                window.speechSynthesis.cancel();
                var msg = new SpeechSynthesisUtterance("{rep_clean}");
                msg.lang = 'fr-FR';
                msg.rate = 1;
                window.speechSynthesis.speak(msg);
            }}
            </script>
        """, height=0)
        st.success(st.session_state.floki_rep)
